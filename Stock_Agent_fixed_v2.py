"""
================================================================================
AI-POWERED STOCK ANALYSIS AGENT (Daily Execution)
================================================================================

PURPOSE:
    Automated daily stock analysis platform that performs dual-layer analysis:
    1. RULE-BASED: Traditional technical + fundamental metrics (for comparison)
    2. AI-BASED: Machine-learning-style composite scoring + recommendations
    
    Keeps both analyses separate to enable AI vs Rule comparison without influence.

ARCHITECTURE:
    - Configuration-driven (config.json for symbols, thresholds, output)
    - Modular data extraction (technical, fundamental, investor)
    - Dual recommendation engines (rule-based vs AI-based)
    - Comprehensive column set for deep analysis
    - MCP-ready for daily scheduled execution

COLUMNS OUTPUT:
    - Identification: Company Name, Full Name, Sector, Market Cap Category
    - Holdings: Fund houses, individuals, average holdings
    - Fundamentals: P/E, ROE, ROCE, Debt/Equity, Growth rates (3Y)
    - Technical: Price, 52W High/Low, RSI, MACD, EMA Status
    - Valuation: Price-to-Book, Price-to-Sales, PEG, Beta
    - Quality: Profit Margin, Operating Margin, Dividend Yield
    - Momentum: RSI trend, MACD trend
    - RULE-BASED: Score, Action, Goal Horizon (for reference)
    - AI-BASED: Score, Ranking, Recommendation, Justification (independent)
    - Metadata: Rank, Last Updated

USAGE:
    python Stock_Agent.py
    (Symbols configured in config.json, not in code)
    
SCHEDULING (MCP Agent - Daily):
    Can be scheduled via cron or Windows Task Scheduler:
    - Daily execution at market close (3:30 PM IST)
    - Output written to AI_STOCK_ANALYSIS.xlsx
    - Logs written to stock_analysis.log

================================================================================
"""

import argparse
import concurrent.futures
import difflib
import sys
from io import StringIO
import yfinance as yf
import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time
import warnings
import re
import json
import logging
import os
import smtplib
import tempfile
from functools import lru_cache

try:
    import yaml
    _YAML_AVAILABLE = True
except ImportError:
    _YAML_AVAILABLE = False
from datetime import datetime, timedelta
from pathlib import Path
from bs4 import BeautifulSoup
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl.styles import PatternFill
try:
    from ta.trend import EMAIndicator, MACD
    from ta.momentum import RSIIndicator
    _TA_AVAILABLE = True
except ImportError:
    _TA_AVAILABLE = False

    class EMAIndicator:
        def __init__(self, close, window=20):
            self.close = pd.Series(close, dtype="float64")
            self.window = int(window)
        def ema_indicator(self):
            return self.close.ewm(span=self.window, adjust=False).mean()

    class MACD:
        def __init__(self, close, window_slow=26, window_fast=12, window_sign=9):
            self.close = pd.Series(close, dtype="float64")
            self.window_slow = int(window_slow)
            self.window_fast = int(window_fast)
            self.window_sign = int(window_sign)
            self._macd = self.close.ewm(span=self.window_fast, adjust=False).mean() - self.close.ewm(span=self.window_slow, adjust=False).mean()
            self._signal = self._macd.ewm(span=self.window_sign, adjust=False).mean()
        def macd(self):
            return self._macd
        def macd_signal(self):
            return self._signal

    class RSIIndicator:
        def __init__(self, close, window=14):
            self.close = pd.Series(close, dtype="float64")
            self.window = int(window)
        def rsi(self):
            delta = self.close.diff()
            gain = delta.clip(lower=0)
            loss = -delta.clip(upper=0)
            avg_gain = gain.ewm(alpha=1 / self.window, min_periods=self.window, adjust=False).mean()
            avg_loss = loss.ewm(alpha=1 / self.window, min_periods=self.window, adjust=False).mean()
            rs = avg_gain / avg_loss.replace(0, pd.NA)
            return 100 - (100 / (1 + rs))

warnings.filterwarnings("ignore")

# =============================
# LOGGING SETUP
# =============================
logging.basicConfig(
    filename="stock_analysis.log",
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Suppress noisy 401/crumb errors from yfinance and related libraries.
# Setting propagate=False prevents sub-loggers (yfinance.base, yfinance.utils, etc.)
# from bubbling up to the root logger which writes to our log file.
for _noisy in ("yfinance", "peewee", "urllib3", "requests"):
    _nl = logging.getLogger(_noisy)
    _nl.setLevel(logging.CRITICAL)
    _nl.propagate = False

# =============================
# CONNECTION POOLING FOR HTTP REQUESTS
# =============================
# Create a session with proper connection pooling and retry strategy.
# This fixes screener.in timeouts by:
# 1. Reusing TCP connections across requests (instead of creating new socket per request)
# 2. Using urllib3's built-in exponential backoff retry (0.3s, 0.6s, 1.2s)
# 3. Limiting concurrent connections to avoid socket exhaustion
SESSION = requests.Session()
adapter = HTTPAdapter(
    pool_connections=10,
    pool_maxsize=20,
    max_retries=Retry(
        total=2,
        connect=2,
        read=2,
        backoff_factor=0.3,
        status_forcelist=[500, 502, 503, 504],
        raise_on_status=False
    )
)
SESSION.mount('http://', adapter)
SESSION.mount('https://', adapter)
SESSION.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})


DEFAULT_TIMEOUT = (5, 20)


def _session_get(url, **kwargs):
    """Centralized HTTP GET with sane defaults and shared session."""
    kwargs.setdefault("timeout", DEFAULT_TIMEOUT)
    return SESSION.get(url, **kwargs)


@lru_cache(maxsize=256)
def _ticker(symbol):
    return yf.Ticker(symbol)


@lru_cache(maxsize=256)
def _ticker_info(symbol):
    try:
        return _ticker(symbol).info or {}
    except Exception as exc:
        logger.debug(f"yfinance info unavailable for {symbol}: {exc}")
        return {}


@lru_cache(maxsize=256)
def _ticker_history(symbol, period="1y", interval="1d"):
    try:
        df = yf.download(symbol, period=period, interval=interval, progress=False, auto_adjust=False, threads=False)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(-1)
        return df.copy()
    except Exception as exc:
        logger.debug(f"yfinance history unavailable for {symbol}: {exc}")
        return pd.DataFrame()


def _atomic_write_text(path_obj, content, encoding="utf-8"):
    path_obj = Path(path_obj)
    path_obj.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", delete=False, encoding=encoding, dir=str(path_obj.parent), suffix=path_obj.suffix or ".tmp") as tmp:
        tmp.write(content)
        tmp_name = tmp.name
    Path(tmp_name).replace(path_obj)
    return path_obj


def _atomic_write_csv(df, path_obj):
    path_obj = Path(path_obj)
    path_obj.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", delete=False, encoding="utf-8", newline="", dir=str(path_obj.parent), suffix=path_obj.suffix or ".csv") as tmp:
        df.to_csv(tmp.name, index=False)
        tmp_name = tmp.name
    Path(tmp_name).replace(path_obj)
    return path_obj

REQUIRED_CONFIG_PATHS = [
    "stocks",
    "symbol_selection.mode",
    "symbol_selection.default_option",
    "symbol_selection.live_rules_file",
    "symbol_selection.live_top_n",
    "symbol_selection.live_min_parameters",
    "symbol_selection.live_min_pass_ratio",
    "symbol_selection.universe_scan_limit",
    "symbol_selection.live_refresh_workers",
    "symbol_selection.fallback_to_config_stocks",
    "ai_suggested_symbols.symbols",
    "ai_suggested_symbols.generated_at",
    "ai_suggested_symbols.rules_file",
    "ai_suggested_symbols.count",
    "export_file",
    "log_file",
    "history_dir",
    "run_report_file",
    "enable_ai_ranking",
    "enable_rule_based",
    "ai_thresholds.strong_buy_min_score",
    "ai_thresholds.buy_min_score",
    "ai_thresholds.hold_min_score",
    "rule_thresholds.rule_score_green_min",
    "rule_thresholds.rule_score_yellow_min",
    "excel_formatting.green_color",
    "excel_formatting.yellow_color",
    "excel_formatting.red_color",
    "parameter_filtering.fast_stage_excluded_keys",
    "parameter_filtering.investor_dependent_keys",
    "parameter_filtering.technical_dependent_keys",
    "display_settings.print_line_width",
    "display_settings.console_header_width",
    "genai.enabled",
    "genai.provider",
    "genai.model",
    "genai.api_key_env",
    "genai.endpoint",
    "genai.temperature",
    "genai.max_tokens",
    "genai.timeout_seconds",
    "email.enabled",
    "email.smtp_host",
    "email.smtp_port",
    "email.use_tls",
    "email.sender_email",
    "email.sender_password_env",
    "email.recipients",
]


def _has_config_path(config, dotted_path):
    node = config
    for part in dotted_path.split("."):
        if not isinstance(node, dict) or part not in node:
            return False
        node = node[part]
    return True


def validate_config(config):
    """Validate required configuration keys are present in config.json."""
    missing = [path for path in REQUIRED_CONFIG_PATHS if not _has_config_path(config, path)]
    if missing:
        raise ValueError(
            "Missing required configuration keys in config.json: " + ", ".join(missing)
        )

# =============================
# CONFIGURATION LOADING
# =============================
def load_config(config_file="config.json"):
    """
    Load configuration from external JSON file.
    
    Args:
        config_file (str): Path to configuration JSON file
        
    Returns:
        dict: Configuration dictionary with stocks, thresholds, output settings
    """
    if not Path(config_file).exists():
        raise FileNotFoundError(
            f"Config file not found: {config_file}. "
            "Create config.json with all required keys before running the agent."
        )

    with open(config_file, 'r', encoding='utf-8') as f:
        config = json.load(f)

    config.setdefault("history_dir", "analysis_history")
    config.setdefault("run_report_file", "latest_run_report.txt")
    config.setdefault("export_file", "AI_STOCK_ANALYSIS.xlsx")
    config.setdefault("log_file", "stock_analysis.log")
    config.setdefault("enable_ai_ranking", True)
    config.setdefault("enable_rule_based", True)

    validate_config(config)
    config["_config_file"] = config_file
    logger.info(f"Loaded configuration from {config_file}")
    return config


def _parse_json_safely(text):
    """Try parsing JSON object from raw LLM text output."""
    if not text:
        return None
    candidate = text.strip()
    try:
        return json.loads(candidate)
    except Exception:
        pass

    start = candidate.find("{")
    end = candidate.rfind("}")
    if start != -1 and end != -1 and end > start:
        snippet = candidate[start:end + 1]
        try:
            return json.loads(snippet)
        except Exception:
            return None
    return None


def call_genai_ghcp(config, system_prompt, user_prompt, max_tokens=None):
    """Call GHCP-compatible chat endpoint and return text response or None on failure."""
    genai_cfg = config.get("genai", {})
    if not genai_cfg.get("enabled", False):
        return None

    provider = str(genai_cfg.get("provider", "")).strip().lower()
    if provider != "ghcp":
        logger.warning(f"GenAI provider '{provider}' is not supported by this integration")
        return None

    api_key_env = genai_cfg.get("api_key_env", "GITHUB_TOKEN")
    api_key = os.getenv(api_key_env, "").strip()
    if not api_key:
        logger.warning(f"GenAI skipped: missing API key env var '{api_key_env}'")
        return None

    endpoint = genai_cfg.get("endpoint", "https://models.inference.ai.azure.com/chat/completions")
    model = genai_cfg.get("model", "gpt-4.1-mini")
    timeout_seconds = int(genai_cfg.get("timeout_seconds", 25))
    temperature = float(genai_cfg.get("temperature", 0.2))
    token_limit = int(max_tokens or genai_cfg.get("max_tokens", 350))

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": token_limit,
    }

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    try:
        response = requests.post(endpoint, headers=headers, json=payload, timeout=timeout_seconds)
        if response.status_code != 200:
            logger.warning(f"GenAI call failed with HTTP {response.status_code}: {response.text[:300]}")
            return None
        body = response.json()
        content = ((body.get("choices") or [{}])[0].get("message") or {}).get("content")
        if content and str(content).strip():
            logger.info(f"GenAI path used via GHCP model '{model}'")
        return content
    except Exception as exc:
        logger.warning(f"GenAI call failed: {exc}")
        return None

# =============================
# SKILL-BASED SCREENER
# =============================

def parse_yaml_front_matter(file_path):
    """
    Extract and parse YAML front-matter (between --- markers) from a Markdown file.
    Requires PyYAML. Raises ImportError if pyyaml is not installed.

    Args:
        file_path (str): Path to the .md file

    Returns:
        dict: Parsed YAML configuration
    """
    if not _YAML_AVAILABLE:
        raise ImportError(
            "PyYAML is required to use the skill-based screener. "
            "Install it with:  pip install pyyaml"
        )
    with open(file_path, 'r', encoding='utf-8') as fh:
        content = fh.read()
    match = re.match(r'^---\n(.*?)\n---', content, re.DOTALL)
    if not match:
        raise ValueError(
            f"No valid YAML front-matter found in '{file_path}'. "
            "The file must start with --- ... --- delimiters."
        )
    return yaml.safe_load(match.group(1))


def load_skill_config(skill_file):
    """
    Load and validate the skill configuration for small/micro cap screening.

    Args:
        skill_file (str): Path to skill markdown file from config.json

    Returns:
        dict: Validated skill config with candidate_universe, filters, scoring_weights
    """
    if not Path(skill_file).exists():
        raise FileNotFoundError(
            f"Skill file '{skill_file}' not found. "
            "Update symbol_selection.live_rules_file in config.json to a valid path."
        )
    config = parse_yaml_front_matter(skill_file)
    if not config.get("candidate_universe"):
        raise ValueError("Skill config must have a non-empty 'candidate_universe' list.")
    logger.info(
        f"[SKILL] Loaded '{config.get('name', 'unnamed')}' v{config.get('version', '?')} "
        f"with {len(config['candidate_universe'])} candidates"
    )
    return config


def _compute_skill_score(roe, revenue_growth, earnings_growth, profit_margin,
                          operating_margin, debt_equity, price_to_book, peg,
                          institutional, weights):
    """
    Compute weighted pre-screen skill score for ranking screened candidates.
    Growth/margin inputs are in decimal form (0â€“1) as returned by yfinance.

    Returns:
        float: Skill pre-screen score
    """
    score = 0.0

    def w(key, default=0):
        return weights.get(key, default)

    # ROE (convert decimal â†’ pct)
    if roe is not None:
        rp = roe * 100
        if rp > 25:   score += w("weight_roe", 15)
        elif rp > 15: score += w("weight_roe", 15) * 0.70
        elif rp > 10: score += w("weight_roe", 15) * 0.40
        elif rp > 5:  score += w("weight_roe", 15) * 0.15

    # Revenue growth
    if revenue_growth is not None:
        rg = revenue_growth * 100
        if rg > 25:   score += w("weight_revenue_growth", 12)
        elif rg > 15: score += w("weight_revenue_growth", 12) * 0.70
        elif rg > 8:  score += w("weight_revenue_growth", 12) * 0.45
        elif rg > 0:  score += w("weight_revenue_growth", 12) * 0.20

    # Earnings growth
    if earnings_growth is not None:
        eg = earnings_growth * 100
        if eg > 25:   score += w("weight_earnings_growth", 12)
        elif eg > 15: score += w("weight_earnings_growth", 12) * 0.70
        elif eg > 8:  score += w("weight_earnings_growth", 12) * 0.45
        elif eg > 0:  score += w("weight_earnings_growth", 12) * 0.20

    # Profit margin
    if profit_margin is not None:
        pm = profit_margin * 100
        if pm > 20:   score += w("weight_profit_margin", 10)
        elif pm > 14: score += w("weight_profit_margin", 10) * 0.75
        elif pm > 8:  score += w("weight_profit_margin", 10) * 0.45
        elif pm > 4:  score += w("weight_profit_margin", 10) * 0.20

    # Debt/Equity (lower is better)
    if debt_equity is not None and debt_equity >= 0:
        if debt_equity < 0.2:   score += w("weight_debt_equity", 10)
        elif debt_equity < 0.6: score += w("weight_debt_equity", 10) * 0.80
        elif debt_equity < 1.0: score += w("weight_debt_equity", 10) * 0.55
        elif debt_equity < 1.5: score += w("weight_debt_equity", 10) * 0.25

    # Price-to-Book (lower is better; skip negative/extreme values)
    if price_to_book is not None and 0.1 < price_to_book < 8:
        pb_score = max(0.0, 1.0 - (price_to_book - 0.5) / 5.5)
        score += w("weight_valuation_pb", 5) * pb_score

    # PEG ratio (lower is better, 0â€“3 range)
    if peg is not None and 0 < peg < 4:
        peg_score = max(0.0, (4 - peg) / 4)
        score += w("weight_peg", 4) * peg_score

    # Institutional holding
    if institutional is not None:
        ip = institutional * 100
        if ip > 30:   score += w("weight_institutional", 5)
        elif ip > 15: score += w("weight_institutional", 5) * 0.65
        elif ip > 5:  score += w("weight_institutional", 5) * 0.35

    return round(score, 2)


def quick_screen_stock(symbol, filters, weights):
    """
    Fast pre-screening for one stock using the yfinance info dict.
    Applies hard filter cut-offs, then computes a skill score for survivors.

    Args:
        symbol (str): NSE stock symbol (e.g. "MCX.NS")
        filters (dict): Filter thresholds from skill config
        weights (dict): Scoring weights from skill config

    Returns:
        dict with {symbol, skill_score, market_cap_cr} if the stock passes,
        or None if it is filtered out or data is unavailable.
    """
    try:
        info = _ticker_info(symbol)
        if not info or not info.get("currentPrice"):
            return None

        current_price    = parse_float(info.get("currentPrice"))
        market_cap_raw   = parse_float(info.get("marketCap"))
        # Convert INR â†’ Crore (1 Cr = 10 million INR)
        market_cap_cr    = round(market_cap_raw / 1e7, 1) if market_cap_raw else None
        pe               = parse_float(info.get("trailingPE"))
        roe              = parse_float(info.get("returnOnEquity"))       # decimal (0â€“1)
        # yfinance returns debtToEquity as 100x for NSE stocks â€” normalise to ratio form.
        _de_raw          = parse_float(info.get("debtToEquity"))
        debt_equity      = _de_raw / 100 if _de_raw is not None else None
        revenue_growth   = parse_float(info.get("revenueGrowth"))        # decimal
        earnings_growth  = parse_float(info.get("earningsGrowth"))       # decimal
        profit_margin    = parse_float(info.get("profitMargins"))        # decimal
        operating_margin = parse_float(info.get("operatingMargins"))     # decimal
        price_to_book    = parse_float(info.get("priceToBook"))
        price_to_sales   = parse_float(info.get("priceToSalesTrailing12Months"))
        peg              = parse_float(info.get("pegRatio"))
        beta             = parse_float(info.get("beta"))
        eps              = parse_float(info.get("epsTrailingTwelveMonths"))
        book_value       = parse_float(info.get("bookValue"))
        institutional    = parse_float(info.get("heldPercentInstitutions"))  # decimal

        f = filters  # short alias

        # â”€â”€ HARD FILTERS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

        # 1. Market cap range
        if market_cap_cr is not None:
            if f.get("min_market_cap_cr") and market_cap_cr < f["min_market_cap_cr"]:
                return None
            if f.get("max_market_cap_cr") and market_cap_cr > f["max_market_cap_cr"]:
                return None

        # 2. Price floor
        if current_price and f.get("min_price_inr") and current_price < f["min_price_inr"]:
            return None

        # 3. PE (only filter positive PE; loss-making stocks have negative PE)
        if pe is not None and pe > 0:
            if f.get("max_pe_ratio") and pe > f["max_pe_ratio"]:
                return None

        # 4. EPS must be positive
        if f.get("min_eps_positive", False) and (eps is None or eps <= 0):
            return None

        # 5. Book value must be positive
        if f.get("min_book_value_positive", False) and (book_value is None or book_value <= 0):
            return None

        # 6. ROE floor
        if roe is not None and f.get("min_roe_pct"):
            if roe * 100 < f["min_roe_pct"]:
                return None

        # 7. Debt/Equity ceiling
        if debt_equity is not None and f.get("max_debt_equity"):
            if debt_equity > f["max_debt_equity"]:
                return None

        # 8. Revenue growth floor
        if revenue_growth is not None and f.get("min_revenue_growth_pct"):
            if revenue_growth * 100 < f["min_revenue_growth_pct"]:
                return None

        # 9. Earnings growth floor
        if earnings_growth is not None and f.get("min_earnings_growth_pct"):
            if earnings_growth * 100 < f["min_earnings_growth_pct"]:
                return None

        # 10. Profit margin floor
        if profit_margin is not None and f.get("min_profit_margin_pct"):
            if profit_margin * 100 < f["min_profit_margin_pct"]:
                return None

        # 11. Beta range
        if beta is not None:
            if f.get("max_beta") and beta > f["max_beta"]:
                return None
            if f.get("min_beta") and beta < f["min_beta"]:
                return None

        # 12. Price-to-Book ceiling
        if price_to_book is not None and price_to_book > 0:
            if f.get("max_price_to_book") and price_to_book > f["max_price_to_book"]:
                return None

        # 13. Price-to-Sales ceiling
        if price_to_sales is not None and price_to_sales > 0:
            if f.get("max_price_to_sales") and price_to_sales > f["max_price_to_sales"]:
                return None

        # 14. PEG ratio ceiling
        if peg is not None and peg > 0:
            if f.get("max_peg_ratio") and peg > f["max_peg_ratio"]:
                return None

        # 15. Institutional holding floor
        if institutional is not None and f.get("min_institutional_holding_pct"):
            if institutional * 100 < f["min_institutional_holding_pct"]:
                return None

        # â”€â”€ COMPUTE SKILL SCORE â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        skill_score = _compute_skill_score(
            roe, revenue_growth, earnings_growth, profit_margin,
            operating_margin, debt_equity, price_to_book, peg,
            institutional, weights
        )
        return {"symbol": symbol, "skill_score": skill_score, "market_cap_cr": market_cap_cr}

    except Exception as e:
        logger.debug(f"[SKILL] Quick screen failed for {symbol}: {e}")
        return None


def discover_symbols_from_skill(skill_config):
    """
    Screen the candidate universe defined in skill_config and return the
    top-ranked symbols for full dual-track analysis.

    Process:
      1. Fetch yfinance info for each candidate
      2. Apply hard filter cut-offs (PE, ROE, Debt, Growth, etc.)
      3. Score survivors using weighted skill scoring
      4. Return the top `output_limit` symbols

    Args:
        skill_config (dict): Loaded skill configuration from load_skill_config()

    Returns:
        list[str]: Top-ranked stock symbols to pass into full analysis
    """
    candidates = skill_config.get("candidate_universe", [])
    filters    = skill_config.get("filters", {})
    weights    = skill_config.get("scoring_weights", {})
    limit      = int(skill_config.get("output_limit", 30))
    min_score  = float(skill_config.get("min_pass_score", 0))

    total = len(candidates)
    logger.info(f"[SKILL] Pre-screening {total} candidates with {len(filters)} filter rules")
    print(f"\n{'='*70}")
    print(f"[SKILL SCREENER] Pre-screening {total} small/micro cap candidates...")
    print(f"{'='*70}")

    passed = []
    for i, symbol in enumerate(candidates, 1):
        print(f"  [{i:>3}/{total}] Screening {symbol:<22}", end=" ", flush=True)
        result = quick_screen_stock(symbol, filters, weights)
        if result and result["skill_score"] >= min_score:
            mcap_str = f"â‚¹{result['market_cap_cr']} Cr" if result["market_cap_cr"] else "mcap N/A"
            print(f"âœ“  score={result['skill_score']:<6}  {mcap_str}")
            passed.append(result)
        else:
            print("âœ—  (filtered out)")
        time.sleep(0.4)

    if not passed:
        logger.warning("[SKILL] No candidates passed filters â€” relaxing to full list with scoring only.")
        print("\nâš  No candidates passed hard filters. Relaxing filters and ranking by skill score.")
        fallback = []
        for symbol in candidates:
            result = quick_screen_stock(symbol, {}, weights)
            if result:
                fallback.append(result)
        fallback.sort(key=lambda x: x["skill_score"], reverse=True)
        top_symbols = [r["symbol"] for r in fallback[:limit]]
    else:
        passed.sort(key=lambda x: x["skill_score"], reverse=True)
        top_symbols = [r["symbol"] for r in passed[:limit]]

    print(f"\n[SKILL] {len(passed)}/{total} stocks passed â†’ top {len(top_symbols)} selected for full analysis.")
    logger.info(f"[SKILL] {len(passed)} passed filters, {len(top_symbols)} selected")
    return top_symbols


def select_run_mode(cli_mode=None):
    """
    Determine the analysis mode from a CLI argument or an interactive prompt.

    Args:
        cli_mode (str | None): Value of --mode CLI argument ('predefined' or 'skill')

    Returns:
        tuple[str, str]: (mode, mode_label)
            mode        â€” 'predefined' or 'skill'
            mode_label  â€” human-readable label used in email subject and reports
    """
    if cli_mode == "predefined":
        return "predefined", "PREDEFINED SYMBOLS"
    if cli_mode == "skill":
        return "skill", "AI LIVE SELECTION"

    # Interactive prompt when no --mode argument is provided
    print("\n" + "=" * 60)
    print("  AI STOCK ANALYSIS AGENT â€” MODE SELECTION")
    print("=" * 60)
    print("  1.  Predefined Symbols   â€” symbols from config.json")
    print("  2.  AI Live Selection    â€” discover best small/micro cap")
    print("                            stocks via rules file in config.json")
    print("=" * 60)
    while True:
        try:
            choice = input("  Enter choice (1 or 2): ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n  Defaulting to Predefined Symbols mode.")
            return "predefined", "PREDEFINED SYMBOLS"
        if choice == "1":
            return "predefined", "PREDEFINED SYMBOLS"
        if choice == "2":
            return "skill", "AI LIVE SELECTION"
        print("  Invalid â€” please enter 1 or 2.")


def normalize_symbol(symbol):
    """Normalize symbol to NSE format expected by downstream analysis."""
    if not symbol:
        return None
    text = str(symbol).strip().upper()
    if not text:
        return None
    if "." not in text:
        text = f"{text}.NS"
    return text


def infer_parameter_keys_from_query(query_lines):
    """Infer known parameter keys from human-readable screening query lines."""
    inferred = []

    for raw_line in query_lines:
        line = raw_line.lower().strip()
        if not line:
            continue

        if "market capitalization" in line and ">" in line and "500" in line:
            inferred.append("market_cap_gt_500")
        if "market capitalization" in line and "<" in line and "50000" in line:
            inferred.append("market_cap_lt_50000")
        if "sales growth" in line and "3" in line and ">" in line and "12" in line:
            inferred.append("sales_growth_3y_gt_12")
        if "profit growth" in line and "3" in line and ">" in line and "12" in line:
            inferred.append("profit_growth_3y_gt_12")
        if "return on equity" in line and ">" in line and "15" in line:
            inferred.append("roe_gt_15")
        if "return on capital employed" in line and ">" in line and "18" in line:
            inferred.append("roce_gt_18")
        if "debt to equity" in line and "<" in line and "0.5" in line:
            inferred.append("debt_lt_0_5")
        if "interest coverage" in line and ">" in line and "4" in line:
            inferred.append("interest_coverage_gt_4")
        if "current ratio" in line and ">" in line and "1.2" in line:
            inferred.append("current_ratio_gt_1_2")
        if "opm" in line and ">" in line and "12" in line:
            inferred.append("operating_margin_gt_12")
        if "peg ratio" in line and "<" in line and "2" in line:
            inferred.append("peg_lt_2_0")
        if "promoter holding" in line and ">" in line and "50" in line:
            inferred.append("promoter_holding_gt_50")
        if "pledged" in line and "<" in line and "3" in line:
            inferred.append("pledged_percentage_lt_3")
        if "cash from operations" in line and ">" in line and "0" in line:
            inferred.append("cash_from_operations_positive")
        if "price to earning" in line and "<" in line and "35" in line:
            inferred.append("pe_lt_35")

    # De-duplicate while preserving order.
    deduped = []
    for key in inferred:
        if key not in deduped:
            deduped.append(key)
    return deduped


def get_cached_ai_symbols(config):
    """Return cached AI-suggested symbols from config if present."""
    cache = config["ai_suggested_symbols"]
    symbols = [normalize_symbol(s) for s in cache.get("symbols", []) if normalize_symbol(s)]
    return symbols


def get_ai_cache_status_line(config):
    """Build a concise status line showing AI symbol cache freshness."""
    cache = config["ai_suggested_symbols"]
    symbols = [normalize_symbol(s) for s in cache.get("symbols", []) if normalize_symbol(s)]
    count = int(cache.get("count") or len(symbols))
    generated_at = cache.get("generated_at") or "never"
    rules_file = cache.get("rules_file") or config["symbol_selection"]["live_rules_file"]

    if count <= 0:
        return f"AI cache status: empty | generated_at={generated_at} | rules={rules_file}"
    return f"AI cache status: {count} symbols | generated_at={generated_at} | rules={rules_file}"


def save_cached_ai_symbols(config, symbols, rules_file):
    """Persist refreshed AI-suggested symbols into config.json."""
    normalized = [normalize_symbol(s) for s in symbols if normalize_symbol(s)]
    payload = {
        "symbols": normalized,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rules_file": rules_file,
        "count": len(normalized),
    }
    config["ai_suggested_symbols"] = payload

    config_path = config.get("_config_file", "config.json")
    path_obj = Path(config_path)
    if not path_obj.exists():
        return

    try:
        existing = json.loads(path_obj.read_text(encoding="utf-8"))
        existing["ai_suggested_symbols"] = payload
        _atomic_write_text(path_obj, json.dumps(existing, indent=2), encoding="utf-8")
    except Exception as exc:
        logger.warning(f"Unable to persist ai_suggested_symbols into {config_path}: {exc}")


def get_fast_stage_keys(parameter_keys, config):
    """Return a subset of keys suitable for quick first-pass screening."""
    excluded = set(config.get("parameter_filtering", {}).get("fast_stage_excluded_keys", []))
    fast_keys = [k for k in parameter_keys if k not in excluded]
    return fast_keys if fast_keys else list(parameter_keys)


def parse_skill_rules_file(skill_file):
    """Read candidate universe, named query parameter lists, and discovery rules from skill markdown.

    Named query sections have the form::

        ## Query: QUERY_NAME
        - parameter_key_one
        - parameter_key_two

    Returns:
        tuple: (symbols, parameter_keys, named_queries, discovery_rules)
            symbols        â€” explicit candidate universe (may be empty)
            parameter_keys â€” flat union of all named-query keys (for backward compat)
            named_queries  â€” dict of {query_name: [key, ...]}
            discovery_rules â€” dict of discovery settings
    """
    path_obj = Path(skill_file)
    if not path_obj.exists():
        raise FileNotFoundError(f"Skill rules file not found: {skill_file}")

    symbols = []
    parameter_keys = []
    named_queries = {}        # {name: [key, ...]}
    discovery_rules = {}
    query_lines = []          # legacy plain-text query lines
    section = None
    current_query_name = None

    for raw_line in path_obj.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        lower_line = line.lower()

        if lower_line.startswith("##"):
            current_query_name = None
            if "candidate universe" in lower_line:
                section = "universe"
            elif "parameter" in lower_line and "key" in lower_line:
                section = "parameters"
            elif "discovery" in lower_line and "rule" in lower_line:
                section = "discovery"
            elif re.match(r"^## query:", lower_line):
                # Named query section: "## Query: SOME_NAME"
                current_query_name = raw_line[raw_line.index(":") + 1:].strip()
                section = "named_query"
                if current_query_name and current_query_name not in named_queries:
                    named_queries[current_query_name] = []
            elif "query" in lower_line:
                section = "query"   # legacy plain-text fallback
            else:
                section = None
            continue

        if section == "universe":
            symbol_match = re.match(r"^-\s*([A-Z0-9][A-Z0-9&.-]{0,20}(?:\.NS|\.BO)?)\s*$", line.upper())
            if symbol_match:
                token = symbol_match.group(1)
                if token not in {"NSE", "BSE"}:
                    symbol = normalize_symbol(token)
                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

        if section == "parameters":
            bullet_match = re.match(r"^-\s*(?:\[[ xX]\]\s*)?([a-z0-9_]+)", lower_line)
            if bullet_match:
                key = bullet_match.group(1).strip()
                if key and key not in parameter_keys:
                    parameter_keys.append(key)

        if section == "discovery":
            kv_match = re.match(r"^-\s*([a-z0-9_]+)\s*:\s*(.+)$", lower_line)
            if kv_match:
                key = kv_match.group(1).strip()
                raw_value = kv_match.group(2).strip()
                if raw_value in {"true", "false"}:
                    discovery_rules[key] = raw_value == "true"
                else:
                    num = parse_float(raw_value)
                    discovery_rules[key] = int(num) if num is not None and float(num).is_integer() else (num if num is not None else raw_value)

        if section == "named_query" and current_query_name is not None:
            bullet_match = re.match(r"^-\s*([a-z0-9_]+)", lower_line)
            if bullet_match:
                key = bullet_match.group(1).strip()
                if key and key not in named_queries[current_query_name]:
                    named_queries[current_query_name].append(key)

        if section == "query":
            clean_line = line.lstrip("- ").strip()
            if clean_line and clean_line.upper() != "AND":
                query_lines.append(clean_line)

    # Build flat parameter_keys: union of named-query keys + directly listed + legacy inferred.
    all_keys = list(parameter_keys)
    for keys in named_queries.values():
        for k in keys:
            if k not in all_keys:
                all_keys.append(k)
    for k in infer_parameter_keys_from_query(query_lines):
        if k not in all_keys:
            all_keys.append(k)
    parameter_keys = all_keys

    return symbols, parameter_keys, named_queries, discovery_rules


def fetch_nse_equity_symbols(limit=350):
    """Fetch NSE equity symbol universe from NSE CSV list."""
    urls = [
        "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
        "https://www1.nseindia.com/content/equities/EQUITY_L.csv",
    ]
    headers = {"User-Agent": "Mozilla/5.0"}

    for url in urls:
        try:
            response = _session_get(url, headers=headers, timeout=15)
            if response.status_code != 200:
                continue
            table = pd.read_csv(StringIO(response.text))
            if "SYMBOL" not in table.columns:
                continue

            symbols = []
            for raw_symbol in table["SYMBOL"].dropna().astype(str):
                symbol = normalize_symbol(raw_symbol)
                if symbol and symbol not in symbols:
                    symbols.append(symbol)

            if limit and len(symbols) > limit:
                # Do not take alphabetically-first tickers only; spread picks across full universe.
                step = len(symbols) / float(limit)
                symbols = [symbols[int(i * step)] for i in range(limit)]
            return symbols
        except Exception:
            continue

    raise ValueError("Unable to fetch NSE equity symbol universe for live discovery")


def parse_market_cap_cr(mcap_value, quote_market_cap):
    """Parse market-cap in crore units using Screener text first, then yfinance fallback."""
    if mcap_value:
        m = re.search(r"([\d,.]+)", str(mcap_value))
        if m:
            parsed = parse_float(m.group(1))
            if parsed is not None:
                return parsed

    if quote_market_cap:
        # yfinance marketCap is usually absolute currency value.
        return round(float(quote_market_cap) / 1e7, 2)

    return None


def build_live_selection_metrics(symbol, parameter_keys=None, config=None, allow_slow_sources=True):
    """Collect normalized metrics used by live AI parameter checks."""
    keys = set(parameter_keys or [])
    investor_dep = set((config or {}).get("parameter_filtering", {}).get("investor_dependent_keys", []))
    technical_dep = set((config or {}).get("parameter_filtering", {}).get("technical_dependent_keys", []))
    needs_technical = bool(keys.intersection(technical_dep))

    technical = None
    if needs_technical:
        if not allow_slow_sources:
            return None
        technical = get_technical(symbol)
        if not technical:
            return None

    if not technical:
        quote_for_price = get_quote_data(symbol)
        price = quote_for_price.get("current_price")
        high_52 = quote_for_price.get("52w_high")
        low_52 = quote_for_price.get("52w_low")
        technical = {
            "price": price,
            "rsi": None,
            "macd": "Data unavailable",
            "ema": "Data unavailable",
            "high": high_52,
            "low": low_52,
        }

    fundamentals = get_fundamentals(symbol) if allow_slow_sources else {}
    quote = get_quote_data(symbol)
    needs_investor_data = bool(keys.intersection(investor_dep))
    investor = get_investor_data(symbol) if (needs_investor_data and allow_slow_sources) else {}

    pe = parse_float(fundamentals.get("pe"))
    roe = parse_float(fundamentals.get("roe"))
    roce = parse_float(fundamentals.get("roce"))
    debt = parse_float(fundamentals.get("debt"))
    debt = debt if debt is not None else quote.get("debt_to_equity")
    interest_coverage = parse_float(fundamentals.get("interest_coverage"))
    if interest_coverage is None:
        interest_coverage = quote.get("interest_coverage")
    current_ratio = parse_float(fundamentals.get("current_ratio"))
    if current_ratio is None:
        current_ratio = quote.get("current_ratio")

    revenue_growth = quote.get("revenue_growth")
    earnings_growth = quote.get("earnings_growth")
    revenue_growth_pct = revenue_growth * 100 if revenue_growth is not None else None
    earnings_growth_pct = earnings_growth * 100 if earnings_growth is not None else None
    sales_growth_3y = parse_float(fundamentals.get("sales_growth_3y"))
    if sales_growth_3y is None:
        sales_growth_3y = revenue_growth_pct
    profit_growth_3y = parse_float(fundamentals.get("profit_growth_3y"))
    if profit_growth_3y is None:
        profit_growth_3y = earnings_growth_pct

    market_cap_cr = parse_market_cap_cr(fundamentals.get("mcap"), quote.get("market_cap"))
    is_micro_cap = market_cap_cr is not None and market_cap_cr < 1000
    is_small_cap = market_cap_cr is not None and 1000 <= market_cap_cr < 5000

    price = technical.get("price")
    high_52w = technical.get("high")
    low_52w = technical.get("low")
    opm_pct = parse_float(fundamentals.get("opm"))
    if opm_pct is None:
        opm_pct = quote.get("operating_margin") * 100 if quote.get("operating_margin") is not None else None
    promoter_holding_pct = parse_float(investor.get("PromoterHolding"))
    if promoter_holding_pct is None:
        promoter_holding_pct = parse_float(fundamentals.get("promoter_holding"))
    pledged_percentage = parse_float(investor.get("PledgedPercentage"))
    if pledged_percentage is None:
        pledged_percentage = parse_float(fundamentals.get("pledged_percentage"))
    cash_from_operations = quote.get("operating_cashflow")

    completeness_fields = [
        pe, roe, roce, debt, quote.get("price_to_book"), quote.get("price_to_sales"),
        quote.get("peg_ratio"), technical.get("rsi"), quote.get("profit_margin"),
        quote.get("operating_margin"), revenue_growth_pct, earnings_growth_pct,
        quote.get("beta"), quote.get("held_percent_institutions"), quote.get("held_percent_insiders"),
        investor.get("MF"), investor.get("FII"), quote.get("dividend_yield"),
        quote.get("eps_trailing_twelve_months"), quote.get("book_value"),
    ]
    available = len([v for v in completeness_fields if v is not None])
    completeness_ratio = available / max(len(completeness_fields), 1)

    return {
        "symbol": symbol,
        "is_micro_cap": is_micro_cap,
        "is_small_cap": is_small_cap,
        "is_not_large_cap": is_micro_cap or is_small_cap,
        "debt": debt,
        "interest_coverage": interest_coverage,
        "current_ratio": current_ratio,
        "roe": roe,
        "roce": roce,
        "pe": pe,
        "pb": quote.get("price_to_book"),
        "ps": quote.get("price_to_sales"),
        "peg": quote.get("peg_ratio"),
        "rsi": technical.get("rsi"),
        "macd_bullish": "bullish" in str(technical.get("macd", "")).lower(),
        "above_ema20": "above 20" in str(technical.get("ema", "")).lower(),
        "above_ema50": "above 50" in str(technical.get("ema", "")).lower(),
        "above_ema200": "above 200" in str(technical.get("ema", "")).lower(),
        "price": price,
        "high_52w": high_52w,
        "low_52w": low_52w,
        "revenue_growth_pct": revenue_growth_pct,
        "earnings_growth_pct": earnings_growth_pct,
        "sales_growth_3y": sales_growth_3y,
        "profit_growth_3y": profit_growth_3y,
        "profit_margin_pct": quote.get("profit_margin") * 100 if quote.get("profit_margin") is not None else None,
        "operating_margin_pct": opm_pct,
        "beta": quote.get("beta"),
        "inst_holding_pct": quote.get("held_percent_institutions") * 100 if quote.get("held_percent_institutions") is not None else None,
        "insider_holding_pct": quote.get("held_percent_insiders") * 100 if quote.get("held_percent_insiders") is not None else None,
        "mf_pct": parse_float(investor.get("MF")),
        "fii_pct": parse_float(investor.get("FII")),
        "promoter_holding_pct": promoter_holding_pct,
        "pledged_percentage": pledged_percentage,
        "cash_from_operations": cash_from_operations,
        "dividend_yield_pct": quote.get("dividend_yield") * 100 if quote.get("dividend_yield") is not None else None,
        "eps": quote.get("eps_trailing_twelve_months"),
        "book_value": quote.get("book_value"),
        "current_price": quote.get("current_price"),
        "market_cap_cr": market_cap_cr,
        "completeness_ratio": completeness_ratio,
    }


def safe_gt(value, threshold):
    return value is not None and value > threshold


def safe_lt(value, threshold):
    return value is not None and value < threshold


# Keys defined here must exactly match what is listed in .github/skill.md query sections.
LIVE_PARAMETER_CHECKS = {
    "market_cap_gt_500":          lambda m: safe_gt(m["market_cap_cr"], 500),
    "market_cap_lt_18000":        lambda m: safe_lt(m["market_cap_cr"], 18000),
    "is_small_or_micro_cap":      lambda m: m["is_not_large_cap"],
    "debt_lt_0_5":                lambda m: safe_lt(m["debt"], 0.5),
    "debt_lt_0_8":                lambda m: safe_lt(m["debt"], 0.8),
    "roe_gt_15":                  lambda m: safe_gt(m["roe"], 15),
    "roce_gt_15":                 lambda m: safe_gt(m["roce"], 15),
    "roce_gt_18":                 lambda m: safe_gt(m["roce"], 18),
    "sales_growth_3y_gt_12":      lambda m: safe_gt(m["sales_growth_3y"], 12),
    "profit_growth_3y_gt_12":     lambda m: safe_gt(m["profit_growth_3y"], 12),
    "operating_margin_gt_12":     lambda m: safe_gt(m["operating_margin_pct"], 12),
    "interest_coverage_gt_4":     lambda m: safe_gt(m["interest_coverage"], 4),
    "cash_from_operations_positive": lambda m: safe_gt(m["cash_from_operations"], 0),
    "promoter_holding_gt_50":     lambda m: safe_gt(m["promoter_holding_pct"], 50),
    "pledged_percentage_lt_3":    lambda m: safe_lt(m["pledged_percentage"], 3),
    "pb_lt_4":                    lambda m: safe_lt(m["pb"], 4),
    "data_completeness_gt_75":    lambda m: safe_gt(m["completeness_ratio"], 0.75),
    "revenue_growth_gt_12":       lambda m: safe_gt(m["revenue_growth_pct"], 12),
    "earnings_growth_gt_12":      lambda m: safe_gt(m["earnings_growth_pct"], 12),
    "peg_lt_1_5":                 lambda m: safe_lt(m["peg"], 1.5),
    "price_above_ema50":          lambda m: m["above_ema50"],
    "price_above_ema200":         lambda m: m["above_ema200"],
    "macd_bullish":               lambda m: m["macd_bullish"],
    "rsi_40_70":                  lambda m: m["rsi"] is not None and 40 <= m["rsi"] <= 70,
    "within_20pct_of_52w_high":   lambda m: m["price"] is not None and m["high_52w"] not in [None, 0] and m["price"] >= (0.8 * m["high_52w"]),
    "pe_lt_35":                   lambda m: safe_lt(m["pe"], 35),
}


def evaluate_live_rules(symbol, parameter_keys, config=None, allow_slow_sources=True):
    """Evaluate parameter-key checks for one symbol and return pass statistics."""
    metrics = build_live_selection_metrics(symbol, parameter_keys, config, allow_slow_sources)
    if not metrics:
        return None

    checks = []
    for key in parameter_keys:
        evaluator = LIVE_PARAMETER_CHECKS.get(key)
        if evaluator is None:
            continue
        checks.append(bool(evaluator(metrics)))

    if not checks:
        return None

    passed = sum(checks)
    total = len(checks)
    return {
        "symbol": symbol,
        "passed": passed,
        "total": total,
        "pass_ratio": passed / total,
        "metrics": metrics,
    }


def _evaluate_best_query(symbol, recognized_named, config):
    """Evaluate a symbol against all 4 named queries; return ALL query scores plus the best one."""
    best = None
    all_query_scores = {}  # Store all query results for transparency
    
    for query_name, keys in recognized_named.items():
        if not keys:
            continue
        result = evaluate_live_rules(symbol, keys, config, True)
        if result is None:
            all_query_scores[query_name] = None
            continue
        result["query_name"] = query_name
        all_query_scores[query_name] = result
        if best is None or result["pass_ratio"] > best["pass_ratio"]:
            best = result
    
    if best:
        best["all_query_scores"] = all_query_scores  # Attach all 4 query scores to best result
    return best


def refresh_ai_suggested_symbols(config):
    """Refresh AI-suggested symbols from named query rules in skill.md and persist into config."""
    selection_cfg = config["symbol_selection"]
    skill_file = selection_cfg["live_rules_file"]

    universe, parameter_keys, named_queries, discovery_rules = parse_skill_rules_file(skill_file)
    min_parameters = int(selection_cfg["live_min_parameters"])

    # Build per-query recognized key lists from LIVE_PARAMETER_CHECKS.
    if named_queries:
        recognized_named = {
            name: [k for k in keys if k in LIVE_PARAMETER_CHECKS]
            for name, keys in named_queries.items()
        }
        max_recognized = max((len(v) for v in recognized_named.values()), default=0)
        if max_recognized < min_parameters:
            raise ValueError(
                f"No skill query has {min_parameters} recognized rules. "
                f"Largest query has {max_recognized} recognized rules."
            )
        logger.info(
            f"Loaded {len(recognized_named)} named queries from {skill_file}: "
            + ", ".join(f"{n}({len(v)}keys)" for n, v in recognized_named.items())
        )
    else:
        # Legacy fallback: single flat key list.
        flat_keys = [k for k in parameter_keys if k in LIVE_PARAMETER_CHECKS]
        if len(flat_keys) < min_parameters:
            raise ValueError(
                f"Skill query maps to {len(flat_keys)} recognized rules. Minimum required: {min_parameters}"
            )
        recognized_named = {"default": flat_keys}
        logger.info(f"Using flat parameter list from {skill_file}: {len(flat_keys)} keys")

    if not universe:
        scan_limit = int(discovery_rules.get("universe_scan_limit", selection_cfg["universe_scan_limit"]))
        universe = fetch_nse_equity_symbols(scan_limit)
        logger.info(
            f"Skill file has no explicit candidate universe. Auto-discovered {len(universe)} NSE symbols (limit={scan_limit})"
        )

    logger.info(f"Refreshing AI suggested symbols from {skill_file}. Candidates={len(universe)}")
    max_selected = int(selection_cfg["live_top_n"])
    min_pass_ratio = float(selection_cfg["live_min_pass_ratio"])
    worker_count = max(1, int(selection_cfg["live_refresh_workers"]))

    # Evaluate each symbol against all named queries; keep the best pass_ratio per symbol.
    full_results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=worker_count) as executor:
        future_to_symbol = {
            executor.submit(_evaluate_best_query, symbol, recognized_named, config): symbol
            for symbol in universe
        }
        for idx, future in enumerate(concurrent.futures.as_completed(future_to_symbol), 1):
            if idx % 20 == 0 or idx == len(universe):
                logger.info(f"Live refresh progress: {idx}/{len(universe)}")
            try:
                result = future.result()
            except Exception as exc:
                symbol = future_to_symbol.get(future, "unknown")
                logger.warning(f"Live refresh evaluation failed for {symbol}: {exc}")
                continue
            if result:
                full_results.append(result)

    if not full_results:
        raise ValueError("No valid candidates could be evaluated for live refresh")

    full_results.sort(key=lambda item: (item["pass_ratio"], item["passed"]), reverse=True)
    shortlisted = [item for item in full_results if item["pass_ratio"] >= min_pass_ratio]
    if not shortlisted:
        shortlisted = full_results[:max_selected]

    symbols = [item["symbol"] for item in shortlisted[:max_selected]]
    if not symbols:
        raise ValueError("Live refresh produced empty symbol list")

    # Preserve query scores for Excel export transparency
    query_scores_map = {item["symbol"]: item.get("all_query_scores", {}) for item in shortlisted[:max_selected]}
    config["_ai_query_scores"] = query_scores_map  # Store for later use in Excel

    save_cached_ai_symbols(config, symbols, skill_file)
    logger.info(f"AI suggested symbols refreshed and cached. Count={len(symbols)}")
    return symbols


def choose_symbol_mode(config):
    """Choose symbol source mode from config or interactive prompt."""
    selection_cfg = config["symbol_selection"]
    mode = str(selection_cfg["mode"]).strip().lower()
    default_option = str(selection_cfg["default_option"]).strip().lower()

    valid_modes = {"predefined", "live_cached", "live_refresh", "live_ai"}

    if mode in {"live_ai", "live"}:
        mode = "live_cached"
    if default_option in {"live_ai", "live"}:
        default_option = "live_cached"

    if mode != "prompt":
        return mode if mode in valid_modes else "predefined"

    try:
        print("\nSelect symbol source mode:")
        print("1) Predefined symbols from config.json")
        print("2) AI suggested symbols (cached list from config.json)")
        print(f"3) Refresh AI suggested symbols now from {selection_cfg['live_rules_file']}")
        prompt_default = {
            "predefined": "1",
            "live_cached": "2",
            "live_refresh": "3",
        }.get(default_option, "1")
        choice = input(f"Enter 1, 2, or 3 (default={prompt_default}): ").strip()
        if not choice:
            choice = prompt_default
        if choice == "2":
            return "live_cached"
        if choice == "3":
            return "live_refresh"
        return "predefined"
    except EOFError:
        return {
            "live_cached": "live_cached",
            "live_refresh": "live_refresh",
        }.get(default_option, "predefined")


def populate_query_scores_for_symbols(symbols, config):
    """Evaluate all symbols against 4 named queries and store scores for Excel transparency."""
    selection_cfg = config["symbol_selection"]
    skill_file = selection_cfg["live_rules_file"]
    
    try:
        universe, parameter_keys, named_queries, discovery_rules = parse_skill_rules_file(skill_file)
        
        if not named_queries:
            # Legacy fallback: single flat key list
            flat_keys = [k for k in parameter_keys if k in LIVE_PARAMETER_CHECKS]
            named_queries = {"default": flat_keys}
        
        # Build recognized key lists per query
        recognized_named = {
            name: [k for k in keys if k in LIVE_PARAMETER_CHECKS]
            for name, keys in named_queries.items()
        }
        
        # Evaluate each symbol against all queries
        query_scores_map = {}
        worker_count = max(1, int(selection_cfg.get("live_refresh_workers", 2)))
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=worker_count) as executor:
            future_to_symbol = {
                executor.submit(_evaluate_best_query, symbol, recognized_named, config): symbol
                for symbol in symbols
            }
            for future in concurrent.futures.as_completed(future_to_symbol):
                try:
                    result = future.result()
                    if result:
                        symbol = result.get("symbol")
                        all_scores = result.get("all_query_scores", {})
                        query_scores_map[symbol] = all_scores
                except Exception:
                    pass
        
        if query_scores_map:
            config["_ai_query_scores"] = query_scores_map
            logger.info(f"Query scores populated for {len(query_scores_map)} symbols")
    except Exception as exc:
        logger.warning(f"Could not populate query scores: {exc}")


def add_query_score_columns(df, config):
    """Add query transparency columns (best query + per-query scores) to analysis dataframe."""
    if df is None or df.empty:
        return df

    query_scores_map = config.get("_ai_query_scores") or {}
    if not isinstance(query_scores_map, dict) or not query_scores_map:
        df["Best Query"] = "Data unavailable"
        df["Best Query Score"] = "Data unavailable"
        return df

    def _normalize_symbol_key(raw):
        text = str(raw or "").strip().upper()
        if not text:
            return ""
        return text if text.endswith(".NS") else f"{text}.NS"

    def _format_score(result):
        if not isinstance(result, dict):
            return "Data unavailable"
        passed = result.get("passed")
        total = result.get("total")
        ratio = result.get("pass_ratio")
        if passed is None or total in (None, 0):
            return "Data unavailable"
        if isinstance(ratio, (int, float)):
            return f"{passed}/{total} ({ratio * 100:.1f}%)"
        return f"{passed}/{total}"

    # Preserve first-seen query order from evaluation map.
    query_names = []
    for scores in query_scores_map.values():
        if isinstance(scores, dict):
            for name in scores.keys():
                if name not in query_names:
                    query_names.append(name)

    if not query_names:
        df["Best Query"] = "Data unavailable"
        df["Best Query Score"] = "Data unavailable"
        return df

    best_query_vals = []
    best_score_vals = []
    per_query_values = {name: [] for name in query_names}

    for _, row in df.iterrows():
        symbol_key = _normalize_symbol_key(row.get("Company Name"))
        all_scores = query_scores_map.get(symbol_key, {})
        if not isinstance(all_scores, dict):
            all_scores = {}

        best_name = "Data unavailable"
        best_ratio = -1.0
        best_result = None

        for name in query_names:
            result = all_scores.get(name)
            per_query_values[name].append(_format_score(result))

            if isinstance(result, dict):
                ratio = result.get("pass_ratio")
                if isinstance(ratio, (int, float)) and ratio > best_ratio:
                    best_ratio = ratio
                    best_name = name
                    best_result = result

        best_query_vals.append(best_name)
        best_score_vals.append(_format_score(best_result))

    df["Best Query"] = best_query_vals
    df["Best Query Score"] = best_score_vals
    for name in query_names:
        df[f"Query Score - {name}"] = per_query_values[name]

    return df


def select_symbols_for_run(config):
    """Select run symbols based on configured mode: predefined, cached AI list, or refresh."""
    mode = choose_symbol_mode(config)
    predefined_stocks = [normalize_symbol(s) for s in config.get("stocks", []) if normalize_symbol(s)]

    if mode == "predefined":
        logger.info(f"Symbol mode=predefined. Using {len(predefined_stocks)} symbols from config.json")
        return predefined_stocks, "PREDEFINED"

    selection_cfg = config["symbol_selection"]
    fallback = bool(selection_cfg["fallback_to_config_stocks"])

    try:
        if mode == "live_cached":
            cached_symbols = get_cached_ai_symbols(config)
            if cached_symbols:
                logger.info(f"Symbol mode=live_cached. Using {len(cached_symbols)} cached AI suggested symbols")
                return cached_symbols, "LIVE_CACHED"

            logger.warning("No cached AI suggested symbols found. Refreshing now.")
            refreshed = refresh_ai_suggested_symbols(config)
            return refreshed, "LIVE_REFRESH"

        if mode == "live_refresh":
            refreshed = refresh_ai_suggested_symbols(config)
            return refreshed, "LIVE_REFRESH"

        # Legacy alias fallback.
        refreshed = refresh_ai_suggested_symbols(config)
        return refreshed, "LIVE_REFRESH"

    except Exception as exc:
        logger.error(f"AI symbol selection failed: {exc}")
        if fallback and predefined_stocks:
            logger.warning(
                f"Falling back to predefined symbols from config.json ({len(predefined_stocks)} symbols)"
            )
            return predefined_stocks, "PREDEFINED"
        raise

# =============================
# UTILITY FUNCTIONS
# =============================

def clean(val):
    """
    Clean and normalize values, replace empty/None with 'Data unavailable'.
    
    Args:
        val: Any value to clean
        
    Returns:
        str or original value: Cleaned value or "Data unavailable"
    """
    if val in ["", None, "-", " "] or pd.isna(val):
        return "Data unavailable"
    return val

def parse_float(val):
    """
    Safely parse float from string or numeric input.
    
    Args:
        val: String or number to parse
        
    Returns:
        float or None: Parsed float or None if parsing fails
    """
    try:
        if val in [None, "", "-", " "]:
            return None
        if isinstance(val, str):
            # Strip INR symbol in both Unicode form (₹) and corrupted latin-1 form (â‚¹)
            val = val.replace(",", "").replace("\u20b9", "").replace("\xe2\x82\xb9", "").replace("â‚¹", "").replace("%", "").strip()
        return float(val)
    except (TypeError, ValueError):
        return None

def normalize_pct(v):
    """
    Normalize percent-like values to percentage points (0-100).
    
    Args:
        v: Decimal (0-1) or percentage string
        
    Returns:
        float: Normalized percentage or None
    """
    num = parse_float(v)
    if num is None:
        return None
    return round(num * 100, 2) if 0 <= num <= 1 else round(num, 2)

def is_institutional_holder(name):
    """
    Classify if a shareholder name is an institutional fund house.
    
    Args:
        name (str): Shareholder name
        
    Returns:
        bool: True if institutional, False otherwise
    """
    if not name:
        return False
    text = name.lower()
    markers = [
        "fund", "mutual", "insurance", "life", "asset", "amc", "trust",
        "institution", "fii", "foreign", "sbi", "hdfc", "icici", "nippon",
        "kotak", "axis", "uti", "aditya birla", "franklin", "invesco"
    ]
    return any(m in text for m in markers)

def is_likely_individual_holder(name):
    """
    Classify if a shareholder name is likely an individual/promoter.
    
    Args:
        name (str): Shareholder name
        
    Returns:
        bool: True if likely individual, False otherwise
    """
    if not name:
        return False
    text = name.strip()
    bad_tokens = ["ltd", "limited", "pvt", "private", "llp", "fund", "trust", "insurance", "asset", "ventures", "capital", "holdings"]
    if any(t in text.lower() for t in bad_tokens):
        return False
    words = re.findall(r"[A-Za-z]+", text)
    return 2 <= len(words) <= 4

def classify_mcap(val):
    """
    Classify market cap into Large/Mid/Small Cap categories.
    
    Args:
        val: Market cap value in Cr.
        
    Returns:
        str: Cap category
    """
    try:
        val = float(str(val).split()[0].replace(",", ""))
        if val > 50000:
            return "Large Cap"
        elif val > 5000:
            return "Mid Cap"
        else:
            return "Small Cap"
    except (TypeError, ValueError, IndexError):
        return "Data unavailable"

# =============================
# DATA EXTRACTION FUNCTIONS
# =============================

def get_company_info(symbol):
    """
    Extract company name and sector from yfinance.
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        tuple: (company_name, sector)
    """
    try:
        info = _ticker_info(symbol)
        return info.get("longName", symbol), info.get("sector", "Data unavailable")
    except:
        return symbol.replace(".NS",""), "Data unavailable"

def get_quote_data(symbol):
    """
    Extract quote-level fundamental and valuation data from yfinance.
    Includes price, growth rates, valuation multiples, margins, and holder percentages.
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        dict: Quote data including valuations, growth, margins, holders
    """
    try:
        info = _ticker_info(symbol)
        return {
            "current_price": parse_float(info.get("currentPrice")),
            "market_cap": parse_float(info.get("marketCap")),
            "52w_high": parse_float(info.get("fiftyTwoWeekHigh")),
            "52w_low": parse_float(info.get("fiftyTwoWeekLow")),
            "revenue_growth": parse_float(info.get("revenueGrowth")),
            "earnings_growth": parse_float(info.get("earningsGrowth")),
            # yfinance returns debtToEquity as a percentage for NSE stocks (e.g. 150 means 1.50).
            # Divide by 100 to convert to the standard ratio form used everywhere else.
            "debt_to_equity": (lambda _de: _de / 100 if _de is not None else None)(parse_float(info.get("debtToEquity"))),
            "industry": info.get("industry", ""),
            "sector_yf": info.get("sector", ""),
            "interest_coverage": parse_float(info.get("interestCoverage")),
            "current_ratio": parse_float(info.get("currentRatio")),
            "operating_cashflow": parse_float(info.get("operatingCashflow")),
            "price_to_book": parse_float(info.get("priceToBook")),
            # yfinance P/S can be inflated for NSE-listed large caps due to a USD/INR
            # currency mismatch (market cap in INR, revenue in USD). Values above 100Ã—
            # are almost certainly wrong for Indian stocks â€” null them out so the cell
            # shows "Data unavailable" rather than a misleading number.
            "price_to_sales": (lambda _ps: _ps if _ps is not None and _ps < 100 else None)(
                parse_float(info.get("priceToSalesTrailing12Months"))),
            "peg_ratio": parse_float(info.get("pegRatio")),
            "held_percent_institutions": parse_float(info.get("heldPercentInstitutions")),
            "held_percent_insiders": parse_float(info.get("heldPercentInsiders")),
            "beta": parse_float(info.get("beta")),
            "dividend_yield": parse_float(info.get("dividendYield")),
            "eps_trailing_twelve_months": parse_float(info.get("epsTrailingTwelveMonths")),
            "book_value": parse_float(info.get("bookValue")),
            "profit_margin": parse_float(info.get("profitMargins")),
            "operating_margin": parse_float(info.get("operatingMargins"))
        }
    except:
        return {k: None for k in [
            "current_price", "market_cap", "52w_high", "52w_low", "revenue_growth", "earnings_growth",
            "debt_to_equity", "interest_coverage", "current_ratio", "operating_cashflow", "price_to_book", "price_to_sales", "peg_ratio",
            "held_percent_institutions", "held_percent_insiders", "beta", "dividend_yield",
            "eps_trailing_twelve_months", "book_value", "profit_margin", "operating_margin",
            "industry", "sector_yf"
        ]}

def resolve_symbol(symbol):
    """
    Attempt to resolve symbol to valid trading format (NSE vs BSE).
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        str or None: Resolved symbol or None if no data available
    """
    candidates = [symbol, symbol.replace(".NS", ".BO")]
    for sym in candidates:
        try:
            df = _ticker_history(sym, period="3mo")
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.droplevel(1)
            if not df.empty:
                return sym
        except Exception as exc:
            logger.debug(f"Symbol resolution failed for {sym}: {exc}")
            continue
    return None

def get_technical(symbol):
    """
    Calculate technical indicators: EMA (20/50/200), RSI, MACD.
    Includes 52W high/low from price history and momentum assessment.
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        dict: Technical indicators and prices, or None if insufficient data
    """
    resolved = resolve_symbol(symbol)
    if not resolved:
        logger.warning(f"[SKIP] No data found for {symbol}")
        return None

    df = yf.download(resolved, period="1y", progress=False)

    # Newer yfinance returns MultiIndex columns (metric, ticker); flatten to single level.
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.droplevel(1)

    if df.empty or len(df) < 100:
        logger.warning(f"[SKIP] Insufficient data for {symbol} - {len(df) if not df.empty else 0} records")
        return None

    close = df["Close"].squeeze()

    df["EMA20"] = EMAIndicator(close, 20).ema_indicator()
    df["EMA50"] = EMAIndicator(close, 50).ema_indicator()
    df["EMA200"] = EMAIndicator(close, 200).ema_indicator()

    df["RSI"] = RSIIndicator(close).rsi()
    df["MACD"] = MACD(close).macd_diff()

    df.dropna(inplace=True)

    if df.empty:
        logger.warning(f"[SKIP] Indicators produced NaN for {symbol}")
        return None

    last = df.iloc[-1]

    price = float(last["Close"])
    rsi = float(last["RSI"])
    macd = float(last["MACD"])

    ema20 = float(last["EMA20"])
    ema50 = float(last["EMA50"])
    ema200 = float(last["EMA200"])

    ema_view = ", ".join([
        "Above 20" if price > ema20 else "Below 20",
        "Above 50" if price > ema50 else "Below 50",
        "Above 200" if price > ema200 else "Below 200"
    ])

    quote = get_quote_data(resolved)

    return {
        "price": round(price, 2),
        "rsi": round(rsi, 2),
        "macd": "Bullish Crossover" if macd > 0 else "Bearish Crossover",
        "ema": ema_view,
        "high": round(quote["52w_high"] if quote["52w_high"] is not None else df["Close"].max(), 2),
        "low": round(quote["52w_low"] if quote["52w_low"] is not None else df["Close"].min(), 2)
    }

def get_fundamentals(symbol):
    """
    Extract fundamental metrics from Screener.in (primary Indian source).
    Covers: P/E, ROE, ROCE, Debt/Equity, Market Cap, Current Price, Book Value,
    Dividend Yield, OPM%, NPM%, EPS, Revenue, Shareholding pattern, and 3Y CAGR.
    All values are in INR â€” no USD/INR currency-mismatch risk.

    Derived ratios computed here:
      price_to_book = Current Price / Book Value  (both INR â†’ ratio is correct)
      price_to_sales = Market Cap (Cr) / Revenue (Cr)  (both INR â†’ ratio is correct)
      npm = Net Profit / Revenue * 100  (from P&L table â†’ correct INR margin)

    Args:
        symbol (str): Stock symbol

    Returns:
        dict: Fundamental metrics with all Indian-sourced values
    """
    sym = symbol.replace(".NS","")
    url = f"https://www.screener.in/company/{sym}/"

    max_retries = 3
    base_delay  = 1

    for attempt in range(max_retries):
        try:
            r = _session_get(url, timeout=10)
            if r.status_code != 200:
                return {}

            soup = BeautifulSoup(r.text, "html.parser")
            data = {}

            # â”€â”€ 1. Ratio box (top summary cards) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            for item in soup.select("li.flex.flex-space-between"):
                name_el = item.find("span", class_="name")
                val_el  = item.find("span", class_="number")
                if not name_el or not val_el:
                    continue
                n = name_el.text.strip()
                v = val_el.text.strip()

                if "Market Cap" in n:
                    data["mcap"] = v
                elif "Current Price" in n:
                    data["current_price_screener"] = v          # INR price from Screener
                elif "High / Low" in n:
                    # Screener shows "High / Low" as a single field; the number cell
                    # contains the 52-week high value (e.g. "2,053")
                    data["high_52w_screener"] = v
                elif "Stock P/E" in n or ("P/E" in n and "Stock" not in data):
                    data["pe"] = v
                elif "Book Value" in n:
                    data["book_value"] = v                       # INR per share
                elif "Dividend Yield" in n:
                    data["dividend_yield"] = v
                elif "ROCE" in n:
                    data["roce"] = v
                elif "ROE" in n:
                    data["roe"] = v
                elif "Face Value" in n:
                    data["face_value"] = v

            # â”€â”€ 2. Tables: growth, P&L, balance sheet, shareholding â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            _pl_annual_done   = False
            _shareholding_done = False
            _growth_section   = None

            for tbl in soup.find_all("table"):
                rows = tbl.find_all("tr")
                if not rows:
                    continue

                # Identify table by its first-row content
                all_first_cells = [c.get_text(" ", strip=True) for c in rows[0].find_all(["th", "td"])]
                first_cell_lower = all_first_cells[0].lower() if all_first_cells else ""
                all_row_labels   = [row.find(["th", "td"]) for row in rows[1:] if row.find(["th", "td"])]
                row_label_texts  = [el.get_text(" ", strip=True).lower() for el in all_row_labels if el]

                # â”€â”€ Compounded Growth tables â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                if "compounded sales" in first_cell_lower or "sales growth" in first_cell_lower:
                    _growth_section = "sales"
                elif "compounded profit" in first_cell_lower or "profit growth" in first_cell_lower:
                    _growth_section = "profit"
                elif "stock price cagr" in first_cell_lower or "return on equity" in first_cell_lower:
                    _growth_section = None  # skip these
                    continue

                if _growth_section in ("sales", "profit"):
                    for row in rows[1:]:
                        cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                        if len(cells) < 2:
                            continue
                        key = cells[0].lower()
                        val = cells[1].replace("%", "").strip()
                        if _growth_section == "sales" and "3" in key:
                            data["sales_growth_3y"] = val
                        elif _growth_section == "profit" and "3" in key:
                            data["profit_growth_3y"] = val
                    _growth_section = None
                    continue

                # â”€â”€ Annual P&L table â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                # Screener layout: first header cell is empty, years start from index 1.
                # Annual: all year cells are "Mar YYYY"; quarterly: contains Jun/Sep/Dec.
                has_pl_rows = any("sales" in t or "net profit" in t or "eps" in t for t in row_label_texts)
                # Check any header cell for year marker (skip the empty first cell)
                any_cell_has_year = any(
                    any(str(y) in cell for y in range(2010, 2028))
                    for cell in all_first_cells
                )
                # Annual tables have only "Mar" months; quarterly has Jun/Sep/Dec too
                has_quarterly_months = any(
                    mon in " ".join(all_first_cells) for mon in ("Jun", "Sep", "Dec")
                )
                is_annual_pl = has_pl_rows and any_cell_has_year and not has_quarterly_months

                if is_annual_pl and not _pl_annual_done:
                    # Column index of most-recent year = last non-empty header cell index
                    # Header: ['', 'Mar 2015', ..., 'Mar 2026']  â†’ index -1 is most recent
                    for row in rows[1:]:
                        cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                        if len(cells) < 2:
                            continue
                        key = cells[0].lower()
                        val = cells[-1]          # most recent year (last column)

                        if key.startswith("sales"):
                            data["revenue_cr"] = val          # Revenue in Cr (INR)
                        elif "net profit" in key:
                            data["net_profit_cr"] = val       # Net Profit in Cr (INR)
                        elif "opm %" in key or key == "opm%":
                            data["opm"] = val                 # Operating Profit Margin %
                        elif "eps in rs" in key or "eps (rs)" in key:
                            data["eps"] = val                 # EPS in INR
                        elif "interest coverage" in key:
                            data["interest_coverage"] = val
                        elif "current ratio" in key:
                            data["current_ratio"] = val

                    _pl_annual_done = True
                    continue

                # â”€â”€ Shareholding table â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                # Detected when rows include "Promoters +", "FIIs +", "DIIs +"
                has_shareholding = any("promoter" in t or "fii" in t or "dii" in t for t in row_label_texts)
                if has_shareholding and not _shareholding_done:
                    # Headers: ['', 'Jun 2023', 'Sep 2023', ...] â€” last 3 quarters = indices -3,-2,-1
                    hdr = [c.get_text(" ", strip=True) for c in rows[0].find_all(["th","td"])]
                    q_labels = [h for h in hdr if h and h != ""][-3:]  # last 3 quarter labels

                    for row in rows[1:]:
                        cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                        if len(cells) < 2:
                            continue
                        key = cells[0].lower()
                        data_cells = [c for c in cells[1:] if c != ""]  # drop empty
                        # Screener shareholding columns are typically oldest -> newest,
                        # so use rightmost values for latest quarters.
                        val_latest = data_cells[-1] if data_cells else ""
                        val_prev = data_cells[-2] if len(data_cells) > 1 else ""
                        val_prev2 = data_cells[-3] if len(data_cells) > 2 else ""

                        if "promoter" in key:
                            data["promoter_holding"] = val_latest.replace("%","").strip()
                        elif "mutual" in key or "mf" in key:
                            data["mf_pct"] = val_latest.replace("%","").strip()
                        elif "fii" in key:
                            data["fii_pct"]  = val_latest.replace("%","").strip()
                            # Build trend string: "Jun25: 10.79% â†’ Sep25: 8.45% â†’ Dec25: 10.79%"
                            if len(q_labels) >= 3 and val_latest and val_prev and val_prev2:
                                data["fii_trend"] = (
                                    f"{q_labels[0] if q_labels else 'Q-3'}: {val_prev2} "
                                    f"> {q_labels[1] if len(q_labels)>1 else 'Q-2'}: {val_prev} "
                                    f"> {q_labels[2] if len(q_labels)>2 else 'Q-1'}: {val_latest}"
                                )
                        elif "dii" in key:
                            data["dii_pct"]  = val_latest.replace("%","").strip()
                            if len(q_labels) >= 3 and val_latest and val_prev and val_prev2:
                                data["dii_trend"] = (
                                    f"{q_labels[0] if q_labels else 'Q-3'}: {val_prev2} "
                                    f"> {q_labels[1] if len(q_labels)>1 else 'Q-2'}: {val_prev} "
                                    f"> {q_labels[2] if len(q_labels)>2 else 'Q-1'}: {val_latest}"
                                )
                        elif "pledged" in key:
                            data["pledged_percentage"] = val_latest.replace("%","").strip()

                    _shareholding_done = True
                    continue

                # â”€â”€ Generic ratio rows (interest coverage, OPM from ratio section) â”€
                for row in rows[1:]:
                    cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                    if len(cells) < 2:
                        continue
                    key = cells[0].lower()
                    val = cells[1]
                    if "interest coverage" in key and "interest_coverage" not in data:
                        data["interest_coverage"] = val
                    elif "current ratio" in key and "current_ratio" not in data:
                        data["current_ratio"] = val
                    elif "opm" in key and "opm" not in data:
                        data["opm"] = val
                    elif ("mutual" in key or "mf" in key) and "mf_pct" not in data:
                        data["mf_pct"] = val.replace("%","").strip()
                    elif "promoter holding" in key and "promoter_holding" not in data:
                        data["promoter_holding"] = val.replace("%","").strip()
                    elif "pledged" in key and "pledged_percentage" not in data:
                        data["pledged_percentage"] = val.replace("%","").strip()

            # Final Screener fallback: parse category percentages from raw shareholding text.
            # Useful when NSE APIs are blocked and table variants change labels/layout.
            if "mf_pct" not in data or "fii_pct" not in data or "dii_pct" not in data:
                share_text = " ".join(soup.get_text(" ", strip=True).split())

                def _extract_pct_from_text(aliases):
                    for alias in aliases:
                        pattern = rf"{alias}[^0-9%]{{0,40}}([0-9]+(?:\\.[0-9]+)?)\\s*%"
                        m = re.search(pattern, share_text, flags=re.IGNORECASE)
                        if m:
                            return m.group(1)
                    return None

                if "mf_pct" not in data:
                    mf_txt = _extract_pct_from_text([r"mutual\\s*funds?", r"mfs?"])
                    if mf_txt is not None:
                        data["mf_pct"] = mf_txt

                if "fii_pct" not in data:
                    fii_txt = _extract_pct_from_text([r"fiis?", r"foreign\\s*institutional\\s*investors?", r"fpi"])
                    if fii_txt is not None:
                        data["fii_pct"] = fii_txt

                if "dii_pct" not in data:
                    dii_txt = _extract_pct_from_text([r"diis?", r"domestic\\s*institutional\\s*investors?"])
                    if dii_txt is not None:
                        data["dii_pct"] = dii_txt

            # â”€â”€ 3a. Scrape sector name from Screener peer breadcrumb â”€â”€â”€â”€â”€â”€â”€â”€â”€
            # The peers section breadcrumb: Broad Sector > Sector > Broad Industry > Industry
            for sec_tag in soup.find_all("section", id="peers"):
                breadcrumb = sec_tag.find("p", class_="sub")
                if breadcrumb:
                    links = [a.get_text(" ", strip=True) for a in breadcrumb.find_all("a")]
                    # links = [BroadSector, Sector, BroadIndustry, Industry]
                    if links:
                        data["screener_sector"]   = links[1] if len(links) > 1 else links[0]
                        data["screener_industry"] = links[-1] if len(links) > 2 else ""

            # â”€â”€ 3. Compute derived ratios (all in INR â€” no currency risk) â”€â”€â”€â”€â”€
            _price = parse_float(str(data.get("current_price_screener", "")).replace(",",""))
            _bv    = parse_float(str(data.get("book_value", "")).replace(",",""))
            _mcap  = parse_float(str(data.get("mcap", "")).replace(",",""))
            _rev   = parse_float(str(data.get("revenue_cr", "")).replace(",",""))
            _np    = parse_float(str(data.get("net_profit_cr", "")).replace(",",""))

            # P/B = Price (INR) / Book Value (INR per share) â€” pure INR ratio
            if _price and _bv and _bv > 0:
                data["price_to_book_screener"] = round(_price / _bv, 2)

            # P/S = Market Cap (Cr) / Revenue (Cr) â€” both INR, no FX mismatch
            if _mcap and _rev and _rev > 0:
                data["price_to_sales_screener"] = round(_mcap / _rev, 2)

            # NPM% = Net Profit / Revenue * 100 â€” from Indian P&L, no FX issue
            if _np is not None and _rev and _rev > 0:
                data["npm"] = round(_np / _rev * 100, 2)

            # PEG (computed from Screener data â€” no yfinance FX risk)
            # PEG = P/E / 3Y Profit Growth CAGR  (both sourced from Screener)
            _pe  = parse_float(str(data.get("pe", "")).replace(",",""))
            _pg3 = parse_float(data.get("profit_growth_3y"))
            if _pe and _pg3 and _pg3 > 0:
                data["peg_computed"] = round(_pe / _pg3, 2)

            return data
            
        except requests.exceptions.RequestException as e:
            # Retry on any network error with backoff (includes Timeout, ConnectionError, etc.)
            if attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt)  # Exponential backoff: 1s, 2s, 4s
                logger.warning(f"Network error for {symbol} (attempt {attempt + 1}/{max_retries}): {type(e).__name__}, retrying in {delay}s...")
                time.sleep(delay)
                continue
            else:
                logger.error(f"Fundamentals extraction failed for {symbol} after {max_retries} attempts: {e}")
                return {}
        except Exception as e:
            logger.error(f"Fundamentals extraction failed for {symbol}: {e}")
            return {}

def get_holder_details(symbol):
    """
    Extract named fund houses and individual shareholders from yfinance.
    Falls back to category-level percentages if named holders unavailable.
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        dict: fund_houses (list), individuals (list), avg_holding (float)
    """
    details = {
        "fund_houses": [],
        "individuals": [],
        "avg_holding": None
    }

    # Source 1: yfinance institutional holders for named fund houses + %.
    try:
        inst_df = yf.Ticker(symbol).institutional_holders
        if inst_df is not None and not inst_df.empty:
            pct_values = []
            for _, row in inst_df.head(5).iterrows():
                holder = str(row.get("Holder", "")).strip()
                pct = normalize_pct(row.get("% Out"))
                if holder:
                    if pct is not None:
                        details["fund_houses"].append(f"{holder} ({pct}%)")
                        pct_values.append(pct)
                    else:
                        details["fund_houses"].append(holder)
            if pct_values:
                details["avg_holding"] = round(sum(pct_values) / len(pct_values), 2)
    except Exception as exc:
        logger.debug(f"Mutual fund holder fetch failed for {symbol}: {exc}")

    # Source 2: yfinance insider roster for prominent individual names.
    try:
        insider_df = _ticker(symbol).insider_roster_holders
        if insider_df is not None and not insider_df.empty and "Name" in insider_df.columns:
            for name in insider_df["Name"].dropna().astype(str).head(5):
                clean_name = name.strip()
                if clean_name and is_likely_individual_holder(clean_name):
                    details["individuals"].append(clean_name)
                    if len(details["individuals"]) >= 3:
                        break
    except Exception as exc:
        logger.debug(f"Insider roster fetch failed for {symbol}: {exc}")

    return details

def _pick_first(d, keys, default=0):
    for k in keys:
        if k in d and d.get(k) not in [None, "", "NA", "N/A"]:
            return d.get(k)
    return default


def _extract_shareholding_from_payload(payload):
    """Extract MF/FII/DII/promoter/pledged percentages from nested NSE payloads."""
    out = {
        "MF": None,
        "FII": None,
        "DII": None,
        "PromoterHolding": None,
        "PledgedPercentage": None,
    }

    def _accept(bucket, value):
        num = parse_float(value)
        if num is None or num < 0 or num > 100:
            return
        if out[bucket] is None:
            out[bucket] = num

    def _walk(node):
        if isinstance(node, list):
            for item in node:
                _walk(item)
            return

        if not isinstance(node, dict):
            return

        # Fast path for common top-level NSE keys.
        _accept("MF", _pick_first(node, [
            "mutualFunds", "mutualFund", "mf", "mfHolding",
            "mutualFundsHolding", "domesticMutualFunds", "mutual_funds", "mf_percent"
        ], None))
        _accept("FII", _pick_first(node, [
            "foreignInstitutions", "foreignInstitutionalInvestors", "fii", "fpi",
            "foreignPortfolioInvestors"
        ], None))
        _accept("DII", _pick_first(node, [
            "domesticInstitutions", "domesticInstitutionalInvestors", "dii",
            "domesticInstitution", "diiHolding"
        ], None))
        _accept("PromoterHolding", _pick_first(node, [
            "promoterHolding", "promoters", "promoter", "promoterShareholding"
        ], None))
        _accept("PledgedPercentage", _pick_first(node, [
            "pledged", "promoterPledged", "pledgedPercentage"
        ], None))

        # Generic key-based fallback for variant payload shapes.
        for key, value in node.items():
            key_norm = str(key).lower().replace(" ", "").replace("_", "")
            if isinstance(value, (dict, list)):
                _walk(value)
                continue

            num = parse_float(value)
            if num is None or num < 0 or num > 100:
                continue

            if ("mutual" in key_norm) or key_norm in ("mf", "mfholding", "mfpercent"):
                _accept("MF", num)
            elif ("foreign" in key_norm and "institution" in key_norm) or key_norm in ("fii", "fpi"):
                _accept("FII", num)
            elif ("domestic" in key_norm and "institution" in key_norm) or key_norm in ("dii", "diiholding"):
                _accept("DII", num)
            elif "promoter" in key_norm and "pledg" not in key_norm:
                _accept("PromoterHolding", num)
            elif "pledg" in key_norm and "non" not in key_norm:
                _accept("PledgedPercentage", num)

    _walk(payload)
    return out

def get_investor_data(symbol):
    """
    Extract category-level shareholding from NSE API: MF %, FII %.
    Uses official NSE API endpoint.
    
    Args:
        symbol (str): Stock symbol
        
    Returns:
        dict: MF percentage, FII percentage
    """
    try:
        base = symbol.replace(".NS", "")
        endpoints = [
            f"https://www.nseindia.com/api/equity-shareholding-pattern?symbol={base}",
            f"https://www.nseindia.com/api/company-share-holding?symbol={base}",
            f"https://www.nseindia.com/api/quote-equity?symbol={base}",
        ]

        s = requests.Session()
        s.headers.update(SESSION.headers)
        s.headers.update({
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": f"https://www.nseindia.com/get-quotes/equity?symbol={base}",
        })

        # Prime cookies/challenge once before API calls.
        s.get("https://www.nseindia.com", timeout=20)
        time.sleep(1)

        parsed = {}
        for url in endpoints:
            try:
                r = s.get(url, timeout=20)
                if r.status_code != 200:
                    continue
                payload = r.json()
                parsed = _extract_shareholding_from_payload(payload)
                if any(parsed.get(k) is not None for k in ("MF", "FII", "DII", "PromoterHolding", "PledgedPercentage")):
                    break
            except Exception as inner_exc:
                logger.debug(f"Investor endpoint failed for {symbol}: {url} -> {inner_exc}")

        if not parsed:
            return {}

        mf = parsed.get("MF")
        fii = parsed.get("FII")
        dii = parsed.get("DII")
        promoter = parsed.get("PromoterHolding")
        pledged = parsed.get("PledgedPercentage")

        mf_num = parse_float(mf) or 0
        fii_num = parse_float(fii) or 0

        return {
            "MF": parse_float(mf),
            "FII": parse_float(fii),
            "DII": parse_float(dii),
            "PromoterHolding": parse_float(promoter),
            "PledgedPercentage": parse_float(pledged),
            "Holding Funds (Top 5)": f"MF: {mf_num}%, FII: {fii_num}%",
            "Avg. Holding %": round((mf_num + fii_num) / 2, 2) if (mf_num or fii_num) else "Data unavailable",
        }

    except Exception as e:
        logger.error(f"Investor data extraction failed for {symbol}: {e}")
        return {}

def format_pct(v):
    """Format percentage value to string."""
    if v is None:
        return "Data unavailable"
    return f"{round(v, 2)}%"


def sanitize_text(value):
    """Normalize bad Excel/encoding artifacts and return ASCII-safe display text."""
    if value is None:
        return value
    text = str(value)

    replacements = {
        "â€”": " - ",
        "—": " - ",
        "–": " - ",
        "â€“": " - ",
        "â€™": "'",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text.encode("ascii", errors="ignore").decode("ascii").strip()

def align_composite_with_ai(final_rating_text, ai_recommendation):
    """
    Strict cap so Composite Rating never exceeds AI action.
    Rebuilds the label deterministically while preserving score block and reasons.
    """
    text = sanitize_text(final_rating_text or "")
    ai_recommendation = sanitize_text(ai_recommendation or "")

    # Expected format:
    # "A+ - STRONG BUY [90/100] High ROE | High ROCE | Strong revenue growth"
    m = re.match(r'^\s*([A-Z][+]?)\s*-\s*([A-Z /]+)\s*(\[\d+/100\])?\s*(.*)$', text)
    if not m:
        # Fallback if format is unexpected
        if ai_recommendation == "STRONG BUY":
            return sanitize_text(text)
        elif ai_recommendation == "BUY":
            return sanitize_text("A - BUY")
        elif ai_recommendation == "HOLD":
            return sanitize_text("B+ - HOLD / WATCH")
        elif ai_recommendation == "REDUCE / AVOID":
            return sanitize_text("C - AVOID / REVIEW")
        return sanitize_text(text)

    grade, old_tag, score_block, reasons = m.groups()
    reasons = sanitize_text(reasons) if reasons else ""

    if ai_recommendation == "STRONG BUY":
        new_grade = "A+"
        new_tag = "STRONG BUY"

    elif ai_recommendation == "BUY":
        new_grade = "A"
        new_tag = "BUY"

    elif ai_recommendation == "HOLD":
        new_grade = "B+"
        new_tag = "HOLD / WATCH"

    elif ai_recommendation == "REDUCE / AVOID":
        new_grade = "C"
        new_tag = "AVOID / REVIEW"

    else:
        # Unknown / N/A -> keep original
        new_grade = grade
        new_tag = old_tag

    rebuilt = f"{new_grade} - {new_tag}"
    if score_block:
        rebuilt += f" {score_block}"
    if reasons:
        rebuilt += f" {reasons}"

    return sanitize_text(rebuilt)

def format_ratio(v, decimals=2, suffix="x"):
    """Format valuation ratios like P/B and P/S correctly."""
    if v is None:
        return "Data unavailable"
    try:
        return f"{round(float(v), decimals)}{suffix}"
    except Exception:
        return "Data unavailable"


def display_or_blank(v):
    """Return display-safe value."""
    if v in [None, "", "Data unavailable"]:
        return "Data unavailable"
    return sanitize_text(v)


def non_empty_flags(flags):
    """Make risk flags explicit instead of blank."""
    if not flags:
        return "None identified"
    if isinstance(flags, str):
        return sanitize_text(flags) if flags.strip() else "None identified"
    text = "; ".join([str(x).strip() for x in flags if str(x).strip()])
    return sanitize_text(text) if text else "None identified"


def format_quarterly_trend(label, values):
    """
    values = [("Sep 2025", 24.84), ("Dec 2025", 25.31), ("Mar 2026", 26.61)]
    """
    cleaned = []
    for period, val in values:
        if val is not None:
            cleaned.append(f"{period}: {round(float(val), 2)}%")
    if not cleaned:
        return f"{label}: Data unavailable"
    return f"{label}: " + " -> ".join(cleaned)


def is_bank_like_business(sector_str="", industry_str=""):
    """
    True only for sectors where leverage is structurally balance-sheet driven:
    banks, NBFCs, insurers, housing finance, lenders.
    Excludes exchanges/capital-market infra names like MCX.
    """
    combined = f"{sector_str or ''} {industry_str or ''}".lower()
    bank_like_tokens = [
        "bank",
        "banks",
        "nbfc",
        "lending",
        "credit",
        "microfinance",
        "housing finance",
        "consumer finance",
        "insurance"
    ]
    exchange_like_exclusions = [
        "exchange",
        "depository",
        "capital markets",
        "market infrastructure",
        "brokerage",
        "asset management"
    ]

    if any(tok in combined for tok in exchange_like_exclusions):
        return False
    return any(tok in combined for tok in bank_like_tokens)


# â”€â”€ Sector / industry median PE lookup (Indian market benchmarks) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_SECTOR_PE_MAP = {
    "information technology": 28, "technology": 28, "software": 26, "it services": 26,
    "banks": 14, "banking": 14, "financial services": 20, "nbfc": 18,
    "insurance": 22, "capital markets": 24, "credit services": 18,
    "consumer defensive": 35, "consumer cyclical": 30, "fmcg": 38, "consumer goods": 35,
    "healthcare": 32, "pharmaceuticals": 30, "pharma": 30,
    "hospitals": 40, "medical care facilities": 40,
    "industrials": 28, "capital goods": 30, "infrastructure": 22,
    "construction": 20, "engineering": 25,
    "energy": 15, "utilities": 14, "oil & gas": 12, "power": 16,
    "basic materials": 18, "metals": 15, "chemicals": 22, "specialty chemicals": 28,
    "auto": 22, "automobiles": 22, "auto components": 20,
    "real estate": 30, "communication services": 20, "telecom": 18,
    "default": 25,"consumer durables": 32,"durables": 32,"home appliances": 34,
    "appliances": 34,"white goods": 34,"electrical equipment": 30,"electricals": 30,
    "cooling products": 30,"air conditioning": 32,"air conditioners": 32,"hvac": 30,"refrigeration": 30
}


def get_sector_pe(sector_str, industry_str=""):
    """Return approximate median sector PE for comparison. Returns (median_pe, label)."""
    s = (sector_str or "").lower().strip()
    i = (industry_str or "").lower().strip()
    for key in _SECTOR_PE_MAP:
        if key in s or key in i:
            return _SECTOR_PE_MAP[key], key.title()
    return _SECTOR_PE_MAP["default"], "Market average"




def get_sector_profile(sector_str="", industry_str=""):
    """Classify sectors for leverage/valuation logic and risk controls."""
    combined = f"{sector_str or ''} {industry_str or ''}".lower()

    financial_tokens = [
        "bank", "financial", "nbfc", "insurance", "credit",
        "capital markets", "asset management", "housing finance",
        "microfinance", "financ"
    ]
    capital_intensive_tokens = [
        "utility", "utilities", "power", "real estate", "construction",
        "infrastructure", "telecom", "hospital", "medical care",
        "shipping", "airline", "cement", "metals"
    ]
    cyclical_tokens = [
        "automobile", "metals", "chemicals", "construction",
        "capital goods", "energy"
    ]

    is_financial = any(tok in combined for tok in financial_tokens)
    is_capital_intensive = any(tok in combined for tok in capital_intensive_tokens)
    is_cyclical = any(tok in combined for tok in cyclical_tokens)
    bank_like = is_bank_like_business(sector_str, industry_str)

    # Debt is NOT directly comparable for pure lenders/insurers;
    # but exchange / capital-market infra names should not be treated as banks.
    debt_metric_applicable = not bank_like
    debt_soft = is_capital_intensive and not bank_like

    if bank_like:
        label = "Financial"
    elif is_capital_intensive:
        label = "Capital Intensive"
    else:
        label = "General"

    return {
        "label": label,
        "is_financial": is_financial,
        "is_bank_like": bank_like,
        "is_capital_intensive": is_capital_intensive,
        "is_cyclical": is_cyclical,
        "debt_metric_applicable": debt_metric_applicable,
        "debt_soft": debt_soft,
    }


def _metric_from_sources(primary, secondary=None, primary_scale=1.0, secondary_scale=1.0, allowed_min=None, allowed_max=None):
    """Prefer primary source, fall back to secondary, and flag large mismatch."""
    p = parse_float(primary)
    s = parse_float(secondary)
    if p is not None:
        p *= primary_scale
    if s is not None:
        s *= secondary_scale
    use = p if p is not None else s
    source = "primary" if p is not None else "secondary" if s is not None else "unavailable"
    conflict = False
    if p is not None and s is not None:
        base = max(abs(p), abs(s), 1.0)
        conflict = abs(p - s) / base > 0.5
    if use is not None and allowed_min is not None and use < allowed_min:
        use = None
    if use is not None and allowed_max is not None and use > allowed_max:
        use = None
    return use, source, conflict


def _label_from_completeness(ratio):
    if ratio >= 0.80:
        return "HIGH"
    if ratio >= 0.60:
        return "MEDIUM"
    return "LOW"


def _build_ai_meta(quote, technical, fundamentals, investor_data):
    sector = fundamentals.get("screener_sector") or fundamentals.get("sector") or quote.get("sector_yf") or ""
    industry = fundamentals.get("screener_industry") or quote.get("industry") or ""
    profile = get_sector_profile(sector, industry)

    roe = parse_float(fundamentals.get("roe"))
    roce = parse_float(fundamentals.get("roce"))
    debt, debt_source, debt_conflict = _metric_from_sources(fundamentals.get("debt"), quote.get("debt_to_equity"), 1.0, 1.0, 0, 25)
    int_cov, _, _ = _metric_from_sources(fundamentals.get("interest_coverage"), quote.get("interest_coverage"), 1.0, 1.0, -50, 1000)
    current_ratio, _, _ = _metric_from_sources(None, quote.get("current_ratio"), 1.0, 1.0, 0, 50)
    cash_from_ops, _, _ = _metric_from_sources(fundamentals.get("cash_from_operations"), quote.get("operating_cashflow"), 1.0, 1.0)
    rev_g_3y = parse_float(fundamentals.get("sales_growth_3y"))
    prof_g_3y = parse_float(fundamentals.get("profit_growth_3y"))
    rev_growth = rev_g_3y if rev_g_3y is not None else (quote.get("revenue_growth") * 100 if quote.get("revenue_growth") is not None else None)
    prof_growth = prof_g_3y if prof_g_3y is not None else (quote.get("earnings_growth") * 100 if quote.get("earnings_growth") is not None else None)
    promoter = parse_float(fundamentals.get("promoter_holding"))
    if promoter is None:
        promoter = parse_float(investor_data.get("PromoterHolding"))
    pledge = parse_float(fundamentals.get("pledged_percentage"))
    if pledge is None:
        pledge = parse_float(investor_data.get("PledgedPercentage"))
    pe = parse_float(fundamentals.get("pe"))
    pb = parse_float(quote.get("price_to_book"))
    ps = parse_float(quote.get("price_to_sales"))
    npm_pct = parse_float(fundamentals.get("npm"))
    if npm_pct is None and quote.get("profit_margin") is not None:
        npm_pct = quote.get("profit_margin") * 100
    opm_pct = parse_float(str(fundamentals.get("opm", "")).replace("%", "").strip()) if fundamentals.get("opm") not in [None, "", "Data unavailable"] else None
    if opm_pct is None and quote.get("operating_margin") is not None:
        opm_pct = quote.get("operating_margin") * 100

    med_pe, pe_label = get_sector_pe(sector, industry)
    pe_ratio = (pe / med_pe) if pe is not None and med_pe else None

    risk_flags = []
    soft_flags = []
    if pledge is not None and pledge > 5:
        risk_flags.append(f"Promoter pledge elevated ({round(pledge, 2)}%)")
    if cash_from_ops is not None and cash_from_ops <= 0:
        risk_flags.append("Negative operating cash flow")
    if profile["debt_metric_applicable"]:
        debt_limit = 1.0 if not profile["debt_soft"] else 2.0
        if debt is not None and debt > debt_limit and (int_cov is None or int_cov < 2.5):
            risk_flags.append(f"Debt stress (D/E {round(debt, 2)})")
    if roe is not None and roe < 8 and (roce is None or roce < 10):
        risk_flags.append("Sub-scale profitability (ROE/ROCE too low)")
    if prof_growth is not None and rev_growth is not None and prof_growth < 0 and rev_growth < 0:
        risk_flags.append("Both revenue and profit growth are negative")
    if pe_ratio is not None and pe_ratio > 2.2 and (prof_growth is None or prof_growth < 15):
        risk_flags.append("Extreme valuation without matching growth")
    if debt_conflict:
        soft_flags.append("Debt mismatch across sources")
    if pe is not None and pb is not None and pe > 50 and pb > 8:
        soft_flags.append("Valuation rich on both PE and PB")
    if current_ratio is not None and current_ratio < 1 and not profile["is_financial"]:
        soft_flags.append("Tight liquidity (current ratio < 1)")
    if npm_pct is not None and npm_pct < 5:
        soft_flags.append("Low net margin")

    completeness_fields = [roe, roce, debt, int_cov, cash_from_ops, rev_growth, prof_growth, promoter, pledge, pe, pb, ps, npm_pct, technical.get("rsi"), technical.get("price")]
    completeness = round(sum(v is not None for v in completeness_fields) / len(completeness_fields), 2)
    if completeness < 0.45:
        risk_flags.append("Low data completeness")

    return {
        "sector": sector,
        "industry": industry,
        "profile": profile,
        "roe": roe,
        "roce": roce,
        "debt": debt,
        "debt_source": debt_source,
        "interest_coverage": int_cov,
        "current_ratio": current_ratio,
        "cash_from_ops": cash_from_ops,
        "revenue_growth_pct": rev_growth,
        "profit_growth_pct": prof_growth,
        "promoter_holding": promoter,
        "pledged_pct": pledge,
        "pe": pe,
        "pb": pb,
        "ps": ps,
        "npm_pct": npm_pct,
        "opm_pct": opm_pct,
        "sector_median_pe": med_pe,
        "sector_pe_label": pe_label,
        "pe_ratio_to_sector": pe_ratio,
        "risk_flags": risk_flags,
        "soft_flags": soft_flags,
        "completeness_ratio": completeness,
        "data_quality_label": _label_from_completeness(completeness),
    }
def get_final_rating(ai_score, rule_score, roe, roce, debt_ratio, profit_growth_3y,
                     revenue_growth_3y, promoter_holding, is_high_de_sector=False):
    """
    Unified Final Rating: combines AI score (70%) + rule score (30%) with quality bonuses.
    Returns a letter grade + summary string.
    """
    composite = round(ai_score * 0.70 + (rule_score / 60 * 100) * 0.30, 1)
    bonus = 0
    reasons = []

    roe_val  = parse_float(roe)  if roe  is not None else None
    roce_val = parse_float(roce) if roce is not None else None
    pg3_val  = parse_float(profit_growth_3y)  if profit_growth_3y  is not None else None
    sg3_val  = parse_float(revenue_growth_3y) if revenue_growth_3y is not None else None
    promo    = parse_float(promoter_holding)  if promoter_holding  is not None else None

    if roe_val  and roe_val  > 20: bonus += 3; reasons.append("High ROE")
    if roce_val and roce_val > 20: bonus += 3; reasons.append("High ROCE")
    if pg3_val  and pg3_val  > 20: bonus += 3; reasons.append("Strong profit growth")
    if sg3_val  and sg3_val  > 15: bonus += 2; reasons.append("Strong revenue growth")
    if promo    and promo    > 50: bonus += 2; reasons.append("High promoter stake")
    if debt_ratio is not None and not is_high_de_sector:
        if debt_ratio > 2.0:  bonus -= 4; reasons.append("High debt")
        elif debt_ratio < 0.3: bonus += 2; reasons.append("Low debt")

    final_score = composite + bonus
    if   final_score >= 80: grade, tag = "A+", "STRONG BUY"
    elif final_score >= 68: grade, tag = "A",  "BUY"
    elif final_score >= 55: grade, tag = "B+", "ACCUMULATE"
    elif final_score >= 42: grade, tag = "B",  "HOLD / WATCH"
    else:                   grade, tag = "C",  "AVOID / REVIEW"

    reason_str = " | ".join(reasons[:3]) if reasons else "Composite score"
    return sanitize_text(f"{grade} - {tag} [{final_score:.0f}/100] {reason_str}")


def load_analysis_sheet(excel_file):
    """Load the Analysis sheet from a generated Excel report."""
    path_obj = Path(excel_file)
    if not path_obj.exists():
        raise FileNotFoundError(
            f"Analysis workbook not found: {excel_file}. Run Stock_Agent.py first to generate it."
        )

    df = pd.read_excel(path_obj, sheet_name="Analysis", keep_default_na=False)
    if df.empty:
        raise ValueError(f"Analysis workbook is empty: {excel_file}")
    return df


def _normalize_stock_lookup(value):
    """Normalize stock lookup text for exact and fuzzy matching."""
    text = str(value or "").strip().upper()
    if not text:
        return ""
    return re.sub(r"[^A-Z0-9]", "", text)


def list_report_stocks(excel_file, limit=None):
    """Return stock names available in the generated Excel report."""
    df = load_analysis_sheet(excel_file)
    names = []
    for raw_name in df.get("Company Name", pd.Series(dtype=str)).astype(str).tolist():
        text = raw_name.strip()
        if text and text not in names:
            names.append(text)
    if limit:
        return names[: max(0, int(limit))]
    return names


def find_stock_report_row(df, stock_name):
    """Find the best-matching stock row from the Analysis sheet."""
    lookup = _normalize_stock_lookup(stock_name)
    if not lookup:
        raise ValueError("Stock name cannot be empty")

    best_row = None
    best_ratio = -1.0
    best_label = None

    for _, row in df.iterrows():
        company_name = str(row.get("Company Name", "")).strip()
        full_name = str(row.get("Company Full Name", "")).strip()
        candidates = [
            company_name,
            normalize_symbol(company_name) if company_name else None,
            full_name,
        ]
        normalized_candidates = [_normalize_stock_lookup(item) for item in candidates if item]
        if not normalized_candidates:
            continue

        if lookup in normalized_candidates:
            return row.to_dict()

        substring_match = any(lookup in item or item in lookup for item in normalized_candidates if item)
        if substring_match:
            ratio = 0.99
        else:
            ratio = max(difflib.SequenceMatcher(None, lookup, item).ratio() for item in normalized_candidates)

        if ratio > best_ratio:
            best_ratio = ratio
            best_row = row.to_dict()
            best_label = company_name or full_name or "unknown"

    if best_row is None or best_ratio < 0.55:
        sample_names = df.get("Company Name", pd.Series(dtype=str)).astype(str).head(10).tolist()
        raise ValueError(
            f"Could not find stock '{stock_name}' in the Analysis sheet. "
            f"Available examples: {', '.join(sample_names)}"
        )

    logger.info(f"Stock lookup '{stock_name}' matched to report row '{best_label}' with ratio={best_ratio:.2f}")
    return best_row


def build_stock_report_snapshot(stock_name, config, excel_file=None):
    """Load a stock row from the generated Excel report and return grounded evidence."""
    workbook = excel_file or config.get("export_file", "AI_STOCK_ANALYSIS.xlsx")
    df = load_analysis_sheet(workbook)
    row = find_stock_report_row(df, stock_name)

    primary_fields = [
        "Company Name",
        "Company Full Name",
        "Sector",
        "Market Cap Category",
        "AI Score",
        "AI Recommendation",
        "AI Confidence",
        "AI Justification",
        "Rule-Based Score",
        "Rule-Based Rating",
        "Rank",
        "Current Price (â‚¹)",
        "PE Ratio",
        "ROE (%)",
        "ROCE (%)",
        "Debt/ Equity",
        "Rev. Growth 3Y (%)",
        "Profit Growth 3Y (%)",
        "Price-to-Book",
        "Price-to-Sales",
        "PEG Ratio (computed)",
        "PEG Ratio",
        "Beta",
        "Dividend Yield (%)",
        "Profit Margin (%)",
        "Operating Margin (%)",
        "Price vs EMA (20/50/200)",
        "RSI (14)",
        "MACD Signal",
        "QUALITY_SMALL_MICRO %",
        "GROWTH_SMALL_MICRO %",
        "MOMENTUM_QUALITY_SMALL_MICRO %",
        "BEST_SMALL_MICRO_CORE %",
        "Last Updated",
    ]

    evidence = {}
    for key in primary_fields:
        value = row.get(key, "")
        if value not in [None, "", "Data unavailable"]:
            evidence[key] = value

    return {
        "excel_file": workbook,
        "stock": row.get("Company Name", stock_name),
        "row": row,
        "evidence": evidence,
    }


def _build_default_stock_answer(question, snapshot):
    """Create a deterministic answer using only workbook evidence."""
    row = snapshot["row"]
    evidence = snapshot["evidence"]
    stock = str(snapshot["stock"])
    ai_recommendation = str(row.get("AI Recommendation", "N/A"))
    ai_score = row.get("AI Score", "N/A")
    ai_confidence = str(row.get("AI Confidence", "N/A"))
    rule_rating = str(row.get("Rule-Based Rating", "N/A"))

    evidence_points = []
    for key in [
        "AI Score",
        "AI Recommendation",
        "AI Confidence",
        "Rule-Based Rating",
        "ROE (%)",
        "ROCE (%)",
        "Debt/ Equity",
        "Rev. Growth 3Y (%)",
        "Profit Growth 3Y (%)",
        "Price vs EMA (20/50/200)",
        "RSI (14)",
        "QUALITY_SMALL_MICRO %",
        "GROWTH_SMALL_MICRO %",
        "MOMENTUM_QUALITY_SMALL_MICRO %",
        "BEST_SMALL_MICRO_CORE %",
    ]:
        if key in evidence:
            evidence_points.append(f"{key}: {evidence[key]}")

    if ai_recommendation in ["STRONG BUY", "BUY"]:
        recommendation = "Positive"
    elif ai_recommendation == "HOLD":
        recommendation = "Neutral"
    elif ai_recommendation == "REDUCE / AVOID":
        recommendation = "Cautious"
    else:
        recommendation = "Insufficient data"

    answer = (
        f"{stock}: AI={ai_recommendation} (score {ai_score}, confidence {ai_confidence}), "
        f"rule view={rule_rating}. This response is based only on the latest generated Excel report."
    )

    caveats = [
        "Model-assisted summary; not personalized financial advice.",
        "Uses the latest generated report snapshot and may not reflect intraday changes.",
    ]

    return {
        "stock": snapshot["stock"],
        "question": str(question).strip(),
        "excel_file": snapshot["excel_file"],
        "recommendation": recommendation,
        "confidence": ai_confidence,
        "answer": answer,
        "evidence_points": evidence_points,
        "caveats": caveats,
        "report_fields": snapshot["evidence"],
    }


def answer_stock_question_from_report(stock_name, question, config, excel_file=None):
    """Answer a stock question using only generated report evidence (deterministic fallback)."""
    snapshot = build_stock_report_snapshot(stock_name, config, excel_file)
    return _build_default_stock_answer(question, snapshot)


def add_score_helpers(df):
    """Add numeric helper columns used for ranking and reporting."""
    helper_df = df.copy()
    ai_series = helper_df["AI Score"] if "AI Score" in helper_df.columns else pd.Series([0] * len(helper_df), index=helper_df.index)
    rule_series = helper_df["Rule-Based Score"] if "Rule-Based Score" in helper_df.columns else pd.Series([0] * len(helper_df), index=helper_df.index)
    rank_series = helper_df["Rank"] if "Rank" in helper_df.columns else pd.Series([99999] * len(helper_df), index=helper_df.index)
    helper_df["AI Score Numeric"] = pd.to_numeric(ai_series, errors="coerce").fillna(0)
    helper_df["Rule Score Numeric"] = pd.to_numeric(rule_series, errors="coerce").fillna(0)
    helper_df["Rank Numeric"] = pd.to_numeric(rank_series, errors="coerce").fillna(99999)

    return helper_df

def parse_numeric_cell(value):
    """Parse numeric values from text cells like '12.5%', '1.2 (high leverage...)', or 'Data unavailable'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"n/a", "na", "none", "nan", "data unavailable"}:
        return None

    text = text.replace(",", "").replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def get_top_ai_recommendations(df, limit=5):
    """Return top AI-ranked recommendations."""
    helper_df = add_score_helpers(df)
    columns = ["Company Name", "AI Score", "AI Recommendation", "AI Confidence", "Rule-Based Rating"]
    available = [c for c in columns if c in helper_df.columns]
    return helper_df.sort_values(by=["AI Score Numeric", "Rank Numeric"], ascending=[False, True])[available].head(limit)


def get_top_rule_recommendations(df, limit=5):
    """Return top rule-based recommendations."""
    helper_df = add_score_helpers(df)
    columns = ["Company Name", "Rule-Based Score", "Rule-Based Rating", "AI Recommendation", "AI Score"]
    available = [c for c in columns if c in helper_df.columns]
    return helper_df.sort_values(by=["Rule Score Numeric", "Rank Numeric"], ascending=[False, True])[available].head(limit)


def normalize_ai_bucket(recommendation):
    """Map AI recommendation labels to comparable buckets."""
    if recommendation in ["STRONG BUY", "BUY"]:
        return "Positive"
    if recommendation == "HOLD":
        return "Neutral"
    if recommendation in ["REDUCE / AVOID", "N/A"]:
        return "Cautious"
    return None


def normalize_rule_bucket(recommendation):
    """Map rule-based labels to comparable buckets."""
    if recommendation == "HOLD / ACCUMULATE":
        return "Positive"
    if recommendation == "HOLD":
        return "Neutral"
    if recommendation == "WATCHLIST":
        return "Cautious"
    return None


def get_common_recommendations(df, limit=5):
    """Return top stocks where AI and rule-based views align by bucket."""
    if df is None or df.empty:
        return pd.DataFrame()

    helper_df = add_score_helpers(df)

    required_cols = ["AI Recommendation", "Rule-Based Rating", "Company Name"]
    missing = [c for c in required_cols if c not in helper_df.columns]
    if missing:
        logger.warning(f"get_common_recommendations skipped because columns are missing: {missing}")
        return pd.DataFrame()

    helper_df["AI Bucket"] = helper_df["AI Recommendation"].apply(normalize_ai_bucket)
    helper_df["Rule Bucket"] = helper_df["Rule-Based Rating"].apply(normalize_rule_bucket)

    aligned = helper_df[
        helper_df["AI Bucket"].notna()
        & helper_df["Rule Bucket"].notna()
        & (helper_df["AI Bucket"] == helper_df["Rule Bucket"])
    ].copy()

    if aligned.empty:
        return pd.DataFrame()

    columns = [
        "Company Name",
        "AI Recommendation",
        "Rule-Based Rating",
        "AI Score",
        "Rule-Based Score",
    ]
    available = [c for c in columns if c in aligned.columns]

    return aligned.sort_values(
        by=["AI Score Numeric", "Rule Score Numeric", "Rank Numeric"],
        ascending=[False, False, True]
    )[available].head(limit)


def select_investor_quality_candidates(df, limit=5):
    """Select fundamentally strong positive AI names for investor-quality shortlist."""
    if df is None or df.empty:
        return pd.DataFrame()

    work = add_score_helpers(df)
    work["ROE Numeric"] = work["ROE (%)"].apply(parse_numeric_cell)
    work["ROCE Numeric"] = work["ROCE (%)"].apply(parse_numeric_cell)
    work["Debt Numeric"] = work["Debt/ Equity"].apply(parse_numeric_cell)
    work["Profit Margin Numeric"] = work["Profit Margin (%)"].apply(parse_numeric_cell)

    positive_mask = work["AI Recommendation"].isin(["STRONG BUY", "BUY"])
    quality_mask = (
        work["ROE Numeric"].fillna(0) >= 15
    ) & (
        work["ROCE Numeric"].fillna(0) >= 15
    ) & (
        work["Debt Numeric"].fillna(999) <= 1.5
    )
    margin_mask = work["Profit Margin Numeric"].fillna(0) >= 8

    filtered = work[positive_mask & quality_mask & margin_mask].copy()
    if filtered.empty:
        filtered = work[positive_mask].copy()

    columns = [
        "Company Name",
        "AI Score",
        "AI Recommendation",
        "ROE (%)",
        "ROCE (%)",
        "Debt/ Equity",
        "Profit Margin (%)",
    ]
    available = [c for c in columns if c in filtered.columns]
    return filtered.sort_values(by=["AI Score Numeric", "Rank Numeric"], ascending=[False, True])[available].head(limit)


def format_dataframe_for_report(df):
    """Render report DataFrames safely for plain-text output."""
    if df is None or df.empty:
        return "No rows available"
    return df.to_string(index=False)


def select_risk_alerts(df, limit=5):
    """Highlight stocks requiring tighter risk monitoring for investors."""
    if df.empty:
        return pd.DataFrame()

    work = add_score_helpers(df)
    work["Debt Numeric"] = work["Debt/ Equity"].apply(parse_numeric_cell)
    work["RSI Numeric"] = work["RSI (14)"].apply(parse_numeric_cell)

    risk_mask = (
        work["AI Recommendation"].isin(["REDUCE / AVOID"]) |
        (work["Debt Numeric"].fillna(0) > 1.8) |
        (work["RSI Numeric"].fillna(50) >= 75)
    )

    filtered = work[risk_mask].copy()
    if filtered.empty:
        return filtered

    filtered["Risk Note"] = filtered.apply(
        lambda row: (
            "AI cautious" if row.get("AI Recommendation") == "REDUCE / AVOID"
            else "Leverage/overheating risk"
        ),
        axis=1,
    )
    columns = [
        "Company Name",
        "AI Recommendation",
        "Rule-Based Rating",
        "Debt/ Equity",
        "RSI (14)",
        "Risk Note",
    ]
    return filtered.sort_values(by=["AI Score Numeric", "Rank Numeric"], ascending=[True, True])[columns].head(limit)


def generate_investor_brief(df, report_sections, total_stocks, run_started_at, config):
    """Create a readable investor-focused summary narrative from model outputs."""
    if df.empty:
        return {
            "headline": "No analyzable stocks in this run.",
            "stance": "No action",
            "key_points": ["Data pipeline returned zero valid rows. Please review sources and symbols."],
            "monitor_points": ["Re-run after data-source health check."],
        }

    work = add_score_helpers(df)
    positive_count = int(work["AI Recommendation"].isin(["STRONG BUY", "BUY"]).sum())
    cautious_count = int((work["AI Recommendation"] == "REDUCE / AVOID").sum())
    hold_count = int((work["AI Recommendation"] == "HOLD").sum())
    positive_ratio = positive_count / max(total_stocks, 1)
    cautious_ratio = cautious_count / max(total_stocks, 1)
    avg_ai_score = round(float(work["AI Score Numeric"].mean()), 1)

    if positive_ratio >= 0.45 and cautious_ratio <= 0.2:
        stance = "Constructive for long-term accumulation"
        action_note = "Prioritize staggered accumulation in quality names over lump-sum entries."
    elif cautious_ratio >= 0.35:
        stance = "Defensive posture for long-term investors"
        action_note = "Prefer capital protection; add only in high-quality names and keep cash buffer."
    else:
        stance = "Selective accumulation"
        action_note = "Accumulate only where quality, valuation, and trend are aligned."

    aligned_count = 0 if report_sections["common_top"].empty else len(report_sections["common_top"])
    headline = (
        f"Investor Brief ({run_started_at.strftime('%d %b %Y')}): {stance}. "
        f"AI average score {avg_ai_score}, with {positive_count}/{total_stocks} positive opportunities."
    )

    key_points = [
        f"Market stance: {stance}.",
        f"Recommendation mix: Positive {positive_count}, Hold {hold_count}, Cautious {cautious_count}.",
        f"AI and rule-based alignment seen in {aligned_count} top names.",
        action_note,
    ]

    monitor_points = [
        "Track any quality candidate whose AI score falls below 60 in consecutive runs.",
        "Review debt-heavy names first if market volatility increases.",
        "Use weekly change section to identify improving or weakening conviction names.",
    ]

    default_brief = {
        "headline": headline,
        "stance": stance,
        "key_points": key_points,
        "monitor_points": monitor_points,
    }

    genai_cfg = config.get("genai", {})
    if not genai_cfg.get("enabled", False):
        logger.info("Investor brief fallback path used: GenAI disabled in config")
        return default_brief

    ai_top_df = report_sections.get("ai_top")
    rule_top_df = report_sections.get("rule_top")
    risk_df = report_sections.get("risk_alerts")

    ai_top_names = []
    if ai_top_df is not None and not ai_top_df.empty and "Company Name" in ai_top_df.columns:
        ai_top_names = ai_top_df["Company Name"].head(5).astype(str).tolist()

    rule_top_names = []
    if rule_top_df is not None and not rule_top_df.empty and "Company Name" in rule_top_df.columns:
        rule_top_names = rule_top_df["Company Name"].head(5).astype(str).tolist()

    risk_names = []
    if risk_df is not None and not risk_df.empty and "Company Name" in risk_df.columns:
        risk_names = risk_df["Company Name"].head(5).astype(str).tolist()

    system_prompt = (
        "You are a cautious long-term investment analyst. "
        "Return ONLY valid JSON with keys: headline, stance, key_points, monitor_points. "
        "key_points and monitor_points must be arrays of 3 short strings each. "
        "Do not provide personalized financial advice."
    )
    user_prompt = json.dumps({
        "date": run_started_at.strftime("%Y-%m-%d"),
        "total_stocks": total_stocks,
        "avg_ai_score": avg_ai_score,
        "positive_count": positive_count,
        "hold_count": hold_count,
        "cautious_count": cautious_count,
        "aligned_count": aligned_count,
        "weekly_summary": report_sections.get("weekly_summary", ""),
        "ai_top_names": ai_top_names,
        "rule_top_names": rule_top_names,
        "risk_names": risk_names,
    })

    raw = call_genai_ghcp(config, system_prompt, user_prompt, max_tokens=420)
    parsed = _parse_json_safely(raw)
    if not isinstance(parsed, dict):
        logger.info("Investor brief fallback path used: GenAI response missing/invalid JSON")
        return default_brief

    headline_out = str(parsed.get("headline", "")).strip()
    stance_out = str(parsed.get("stance", "")).strip()
    key_out = parsed.get("key_points", [])
    monitor_out = parsed.get("monitor_points", [])

    if not headline_out or not stance_out:
        logger.info("Investor brief fallback path used: GenAI output missing headline/stance")
        return default_brief
    if not isinstance(key_out, list) or not isinstance(monitor_out, list):
        logger.info("Investor brief fallback path used: GenAI output malformed key_points/monitor_points")
        return default_brief

    logger.info("Investor brief GenAI path used")

    return {
        "headline": headline_out,
        "stance": stance_out,
        "key_points": [str(x) for x in key_out[:3]] or default_brief["key_points"],
        "monitor_points": [str(x) for x in monitor_out[:3]] or default_brief["monitor_points"],
    }


def build_text_investor_report(report_sections, export_file, total_stocks, run_started_at):
    """Build a structured plain-text investor report."""
    brief = report_sections["investor_brief"]
    mode_label = report_sections.get("mode_label", "PREDEFINED SYMBOLS")
    run_meta = pd.DataFrame([
        {
            "Headline": brief["headline"],
            "Mode": mode_label,
            "Run timestamp": run_started_at.strftime('%Y-%m-%d %H:%M:%S'),
            "Total stocks analyzed": total_stocks,
            "Excel export": export_file,
        }
    ])
    key_insights_df = pd.DataFrame(
        [{"Insight": point} for point in brief["key_points"]]
    )
    monitor_df = pd.DataFrame(
        [{"What to monitor next": point} for point in brief["monitor_points"]]
    )
    weekly_df = pd.DataFrame([
        {"Weekly change summary": report_sections["weekly_summary"]}
    ])

    lines = [
        "1) Run overview",
        format_dataframe_for_report(run_meta),
        "",
        "2) Key investor insights",
        format_dataframe_for_report(key_insights_df),
        "",
        "3) Top long-term quality candidates",
        format_dataframe_for_report(report_sections["quality_top"]),
        "",
        "4) Top 5 AI recommendations",
        format_dataframe_for_report(report_sections["ai_top"]),
        "",
        "5) Top 5 rule-based recommendations",
        format_dataframe_for_report(report_sections["rule_top"]),
        "",
        "6) AI + Rule aligned names (high-conviction overlap)",
        format_dataframe_for_report(report_sections["common_top"]),
        "",
        "7) Risk alerts",
        format_dataframe_for_report(report_sections["risk_alerts"]),
        "",
        "8) Weekly change summary",
        format_dataframe_for_report(weekly_df),
        "",
        "9) What to monitor next",
        format_dataframe_for_report(monitor_df),
    ]
    lines.extend([
        "",
        "Note: This is a model-assisted investor brief and not personalized financial advice.",
    ])
    return "\n".join(lines)


def _safe_float(value):
    """Best-effort conversion for values like '12.4%' or 'Data unavailable'."""
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip().replace("%", "").replace(",", "")
    if not text or text.lower() in {"n/a", "nan", "none", "data unavailable"}:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _build_rule_justification(row):
    """Create rule-based rationale from key technical and fundamental signals."""
    reasons = []

    rule_score = _safe_float(row.get("Rule-Based Score"))
    debt_equity = _safe_float(row.get("Debt/ Equity"))
    roe = _safe_float(row.get("ROE (%)"))
    roce = _safe_float(row.get("ROCE (%)"))
    rev_growth = _safe_float(row.get("Rev. Growth 3Y (%)"))
    profit_growth = _safe_float(row.get("Profit Growth 3Y (%)"))
    rsi = _safe_float(row.get("RSI (14)"))
    macd = str(row.get("MACD Signal", ""))
    ema = str(row.get("Price vs EMA (20/50/200)", ""))

    if debt_equity is not None and debt_equity <= 0.5:
        reasons.append(f"low leverage (Debt/Equity={debt_equity})")
    if roe is not None and roe >= 15:
        reasons.append(f"healthy ROE ({roe}%)")
    if roce is not None and roce >= 15:
        reasons.append(f"strong ROCE ({roce}%)")
    if rev_growth is not None and rev_growth >= 10:
        reasons.append(f"good revenue growth ({rev_growth}% over 3Y)")
    if profit_growth is not None and profit_growth >= 10:
        reasons.append(f"good profit growth ({profit_growth}% over 3Y)")
    if rsi is not None and 40 <= rsi <= 70:
        reasons.append(f"balanced RSI ({rsi})")
    if "bull" in macd.lower() or "positive" in macd.lower():
        reasons.append(f"positive MACD signal ({macd})")
    if "above" in ema.lower() or "bull" in ema.lower():
        reasons.append(f"price/EMA trend supportive ({ema})")

    if not reasons:
        reasons.append("limited strong rule triggers; score driven by mixed available signals")

    if rule_score is not None:
        score_note = f"Rule score {round(rule_score, 2)}"
    else:
        score_note = "Rule score unavailable"

    return f"{score_note}. Key drivers: " + "; ".join(reasons) + "."


def build_score_justification_sheet(df):
    """Build a detailed justification sheet for AI and rule-based ranking decisions."""
    if df is None or df.empty:
        return pd.DataFrame()

    work = add_score_helpers(df)
    work["AI Rank"] = work["AI Score Numeric"].rank(method="min", ascending=False).astype(int)
    work["Rule Rank"] = work["Rule Score Numeric"].rank(method="min", ascending=False).astype(int)

    rows = []
    for _, row in work.iterrows():
        ai_score = _safe_float(row.get("AI Score"))
        ai_rec = row.get("AI Recommendation", "N/A")

        if ai_score is None:
            selection_summary = "AI score unavailable; relied on rule/available metrics"
        elif ai_score >= 75:
            selection_summary = "High-priority candidate driven by strong AI score and supportive fundamentals"
        elif ai_score >= 60:
            selection_summary = "Good candidate with positive AI signal and acceptable risk profile"
        elif ai_score >= 45:
            selection_summary = "Watchlist candidate with mixed signal quality"
        else:
            selection_summary = "Lower conviction candidate; needs stronger trend/fundamental confirmation"

        rows.append({
            "Overall Rank": row.get("Rank"),
            "Company Name": row.get("Company Name"),
            "AI Rank": int(row.get("AI Rank", 0)),
            "AI Score": row.get("AI Score"),
            "AI Recommendation": ai_rec,
            "AI Confidence": row.get("AI Confidence", "N/A"),
            "Rule Rank": int(row.get("Rule Rank", 0)),
            "Rule-Based Score": row.get("Rule-Based Score"),
            "Rule-Based Rating": row.get("Rule-Based Rating", "N/A"),
            "Selection Summary": selection_summary,
            "AI Detailed Justification": row.get("AI Justification", "AI justification unavailable"),
            "Rule Detailed Justification": _build_rule_justification(row),
        })

    return pd.DataFrame(rows)


def apply_score_justification_colors(writer, config):
    """Apply green/yellow/red fills to AI and rule score columns on justification sheet."""
    ws = writer.sheets.get("Score_Justification")
    if ws is None or ws.max_row < 2:
        return

    header_to_col = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}
    ai_col = header_to_col.get("AI Score")
    rule_col = header_to_col.get("Rule-Based Score")
    if not ai_col and not rule_col:
        return

    excel_fmt = config.get("excel_formatting", {})
    green_fill = PatternFill(fill_type="solid", fgColor=excel_fmt.get("green_color", "C6EFCE"))
    yellow_fill = PatternFill(fill_type="solid", fgColor=excel_fmt.get("yellow_color", "FFEB9C"))
    red_fill = PatternFill(fill_type="solid", fgColor=excel_fmt.get("red_color", "FFC7CE"))

    ai_cfg = config.get("ai_thresholds", {})
    ai_green_min = float(ai_cfg.get("strong_buy_min_score", 75))
    ai_yellow_min = float(ai_cfg.get("buy_min_score", 60))
    
    rule_cfg = config.get("rule_thresholds", {})
    rule_green_min = float(rule_cfg.get("rule_score_green_min", 70.0))
    rule_yellow_min = float(rule_cfg.get("rule_score_yellow_min", 55.0))

    for row in range(2, ws.max_row + 1):
        if ai_col:
            ai_value = _safe_float(ws.cell(row=row, column=ai_col).value)
            if ai_value is not None:
                if ai_value >= ai_green_min:
                    ws.cell(row=row, column=ai_col).fill = green_fill
                elif ai_value >= ai_yellow_min:
                    ws.cell(row=row, column=ai_col).fill = yellow_fill
                else:
                    ws.cell(row=row, column=ai_col).fill = red_fill

        if rule_col:
            rule_value = _safe_float(ws.cell(row=row, column=rule_col).value)
            if rule_value is not None:
                if rule_value >= rule_green_min:
                    ws.cell(row=row, column=rule_col).fill = green_fill
                elif rule_value >= rule_yellow_min:
                    ws.cell(row=row, column=rule_col).fill = yellow_fill
                else:
                    ws.cell(row=row, column=rule_col).fill = red_fill


def dataframe_to_html_table(df):
    """Render compact HTML table for email reports."""
    if df is None or df.empty:
        return "<p><em>No rows available</em></p>"
    return df.to_html(index=False, border=0, justify="left")


def build_html_investor_report(report_sections, export_file, total_stocks, run_started_at):
    """Build a readable HTML investor report for email clients."""
    brief = report_sections["investor_brief"]
    mode_label = report_sections.get("mode_label", "PREDEFINED SYMBOLS")
    run_meta = pd.DataFrame([
        {
            "Headline": brief["headline"],
            "Mode": mode_label,
            "Run timestamp": run_started_at.strftime('%Y-%m-%d %H:%M:%S'),
            "Total stocks analyzed": total_stocks,
            "Excel export": export_file,
        }
    ])
    key_insights_df = pd.DataFrame(
        [{"Insight": point} for point in brief["key_points"]]
    )
    monitor_df = pd.DataFrame(
        [{"What to monitor next": point} for point in brief["monitor_points"]]
    )
    weekly_df = pd.DataFrame([
        {"Weekly change summary": report_sections["weekly_summary"]}
    ])

    return f"""
<html>
  <body style=\"font-family:Segoe UI,Arial,sans-serif;line-height:1.45;color:#1f2937;\">
    <h2 style=\"margin-bottom:8px;\">Daily Investor Brief</h2>

    <h3>1) Run overview</h3>
    {dataframe_to_html_table(run_meta)}

    <h3>1) Key investor insights</h3>
    {dataframe_to_html_table(key_insights_df)}

    <h3>2) Top long-term quality candidates</h3>
    {dataframe_to_html_table(report_sections['quality_top'])}

    <h3>3) Top 5 AI recommendations</h3>
    {dataframe_to_html_table(report_sections['ai_top'])}

    <h3>4) Top 5 rule-based recommendations</h3>
    {dataframe_to_html_table(report_sections['rule_top'])}

    <h3>5) AI + Rule aligned names</h3>
    {dataframe_to_html_table(report_sections['common_top'])}

    <h3>6) Risk alerts</h3>
    {dataframe_to_html_table(report_sections['risk_alerts'])}

    <h3>7) Weekly change summary</h3>
        {dataframe_to_html_table(weekly_df)}

    <h3>8) What to monitor next</h3>
        {dataframe_to_html_table(monitor_df)}

    <p style=\"color:#6b7280;font-size:12px;\">This is a model-assisted investor brief and not personalized financial advice.</p>
  </body>
</html>
"""


def build_compact_fallback_summary(report_sections, export_file, total_stocks, run_started_at):
    """Build a short plain-text fallback for email clients that block HTML."""
    brief = report_sections["investor_brief"]

    def top_names(df):
        if df is None or df.empty or "Company Name" not in df.columns:
            return "None"
        return ", ".join(df["Company Name"].head(3).astype(str).tolist())

    ai_top_names = top_names(report_sections.get("ai_top"))
    rule_top_names = top_names(report_sections.get("rule_top"))
    common_count = 0 if report_sections.get("common_top") is None or report_sections["common_top"].empty else len(report_sections["common_top"])

    lines = [
        f"Daily Investor Brief - {run_started_at.strftime('%Y-%m-%d')}",
        "",
        f"Headline: {brief['headline']}",
        f"Total stocks analyzed: {total_stocks}",
        f"Top AI names: {ai_top_names}",
        f"Top rule-based names: {rule_top_names}",
        f"AI + Rule aligned names count: {common_count}",
        f"Weekly summary: {report_sections['weekly_summary']}",
        "",
        f"Detailed tabular report is available in the HTML section and attachment: {export_file}",
    ]
    return "\n".join(lines)


def parse_snapshot_timestamp(path_obj):
    """Parse archived snapshot timestamp from file name."""
    stem = path_obj.stem
    prefix = "analysis_"
    if not stem.startswith(prefix):
        return None
    raw_value = stem[len(prefix):]
    for pattern in ["%Y%m%d_%H%M%S", "%Y%m%d"]:
        try:
            return datetime.strptime(raw_value, pattern)
        except ValueError:
            continue
    return None


def load_comparison_snapshot(history_dir, run_started_at):
    """Load the most relevant historical snapshot for weekly comparison."""
    history_path = Path(history_dir)
    if not history_path.exists():
        return None, None

    snapshots = []
    for file_path in history_path.glob("analysis_*.csv"):
        timestamp = parse_snapshot_timestamp(file_path)
        if timestamp and timestamp < run_started_at:
            snapshots.append((timestamp, file_path))

    if not snapshots:
        return None, None

    snapshots.sort(key=lambda item: item[0])
    weekly_target = run_started_at - timedelta(days=7)
    weekly_candidates = [item for item in snapshots if item[0] <= weekly_target]

    if weekly_candidates:
        selected_timestamp, selected_path = weekly_candidates[-1]
        label = f"vs snapshot from {selected_timestamp.strftime('%Y-%m-%d')}"
    else:
        selected_timestamp, selected_path = snapshots[-1]
        label = f"vs previous available snapshot from {selected_timestamp.strftime('%Y-%m-%d')}"

    try:
        return pd.read_csv(selected_path), label
    except Exception as exc:
        logger.warning(f"Unable to load comparison snapshot {selected_path}: {exc}")
        return None, None


def build_change_list(current_names, previous_names):
    """Return a readable added/dropped summary."""
    added = sorted(set(current_names) - set(previous_names))
    dropped = sorted(set(previous_names) - set(current_names))
    added_text = ", ".join(added) if added else "none"
    dropped_text = ", ".join(dropped) if dropped else "none"
    return added_text, dropped_text


def generate_weekly_change_summary(current_df, previous_df, comparison_label):
    """Generate a concise week-over-week change summary for the email report."""
    if previous_df is None or previous_df.empty:
        return (
            "Weekly summary unavailable because no prior archived snapshot exists yet. "
            "Future runs will compare against files stored in the history directory."
        )

    current_ai = get_top_ai_recommendations(current_df, limit=5)["Company Name"].tolist()
    previous_ai = get_top_ai_recommendations(previous_df, limit=5)["Company Name"].tolist()
    current_rule = get_top_rule_recommendations(current_df, limit=5)["Company Name"].tolist()
    previous_rule = get_top_rule_recommendations(previous_df, limit=5)["Company Name"].tolist()
    current_common = get_common_recommendations(current_df, limit=5)
    previous_common = get_common_recommendations(previous_df, limit=5)

    ai_added, ai_dropped = build_change_list(current_ai, previous_ai)
    rule_added, rule_dropped = build_change_list(current_rule, previous_rule)
    common_added, common_dropped = build_change_list(
        current_common["Company Name"].tolist() if not current_common.empty else [],
        previous_common["Company Name"].tolist() if not previous_common.empty else [],
    )

    current_ai_counts = current_df["AI Recommendation"].value_counts()
    previous_ai_counts = previous_df["AI Recommendation"].value_counts()
    all_labels = sorted(set(current_ai_counts.index).union(set(previous_ai_counts.index)))
    distribution_changes = []
    for label in all_labels:
        delta = int(current_ai_counts.get(label, 0) - previous_ai_counts.get(label, 0))
        if delta:
            direction = "+" if delta > 0 else ""
            distribution_changes.append(f"{label} {direction}{delta}")
    distribution_text = "; ".join(distribution_changes) if distribution_changes else "No AI recommendation distribution changes"

    return " ".join([
        f"Comparison basis: {comparison_label}.",
        f"AI top 5 added: {ai_added}; dropped: {ai_dropped}.",
        f"Rule-based top 5 added: {rule_added}; dropped: {rule_dropped}.",
        f"Aligned AI+Rule names added: {common_added}; dropped: {common_dropped}.",
        f"AI recommendation mix change: {distribution_text}.",
    ])


def archive_analysis_snapshot(df, history_dir, run_started_at):
    """Persist a point-in-time CSV snapshot for later weekly comparisons."""
    history_path = Path(history_dir)
    history_path.mkdir(parents=True, exist_ok=True)
    archive_path = history_path / f"analysis_{run_started_at.strftime('%Y%m%d_%H%M%S')}.csv"
    _atomic_write_csv(df, archive_path)
    return archive_path


def build_run_report(df, comparison_df, comparison_label, config):
    """Assemble all report sections used for text output and email."""
    ai_top = get_top_ai_recommendations(df, limit=5)
    rule_top = get_top_rule_recommendations(df, limit=5)
    common_top = get_common_recommendations(df, limit=5)
    quality_top = select_investor_quality_candidates(df, limit=5)
    risk_alerts = select_risk_alerts(df, limit=5)
    weekly_summary = generate_weekly_change_summary(df, comparison_df, comparison_label)
    investor_brief = generate_investor_brief(df, {
        "common_top": common_top,
        "ai_top": ai_top,
        "rule_top": rule_top,
        "risk_alerts": risk_alerts,
        "weekly_summary": weekly_summary,
    }, len(df), datetime.now(), config)
    return {
        "ai_top": ai_top,
        "rule_top": rule_top,
        "common_top": common_top,
        "quality_top": quality_top,
        "risk_alerts": risk_alerts,
        "weekly_summary": weekly_summary,
        "investor_brief": investor_brief,
    }


def write_run_report(report_file, report_sections, export_file, total_stocks, run_started_at):
    """Write the latest run summary to a text file."""
    report_text = build_text_investor_report(report_sections, export_file, total_stocks, run_started_at)
    _atomic_write_text(Path(report_file), report_text, encoding="utf-8")
    return report_text


def send_email_report(config, report_sections, export_file, total_stocks, run_started_at, mode_label="PREDEFINED SYMBOLS"):
    """Email the latest stock-analysis summary when SMTP is configured.

    Args:
        mode_label (str): 'PREDEFINED SYMBOLS' or 'AI LIVE SELECTION' â€” used in subject line.
    """
    email_cfg = config.get("email", {})
    if not email_cfg.get("enabled"):
        return False, "Email disabled in config"

    password_env_key = email_cfg.get("sender_password_env", "STOCK_AGENT_EMAIL_PASSWORD")
    sender_password = os.getenv(password_env_key, "")

    required_fields = [
        "smtp_host",
        "smtp_port",
        "sender_email",
    ]
    missing = [field for field in required_fields if not email_cfg.get(field)]
    recipients = email_cfg.get("recipients", [])
    if not sender_password:
        missing.append(f"env:{password_env_key}")

    if missing or not recipients:
        missing_parts = ", ".join(missing + (["recipients"] if not recipients else []))
        return False, f"Email skipped because config is incomplete: {missing_parts}"

    # Email subject includes the mode label so recipients immediately know
    # whether this run used predefined symbols or AI-driven live selection.
    subject = (
        f"Investor Brief | {mode_label} | Daily Stock Analysis"
        f" | {run_started_at.strftime('%Y-%m-%d')}"
    )
    fallback_body = build_compact_fallback_summary(report_sections, export_file, total_stocks, run_started_at)
    html_body = build_html_investor_report(report_sections, export_file, total_stocks, run_started_at)

    message = MIMEMultipart("mixed")
    message["From"] = email_cfg["sender_email"]
    message["To"] = ", ".join(recipients)
    message["Subject"] = subject

    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(fallback_body, "plain", "utf-8"))
    alternative.attach(MIMEText(html_body, "html", "utf-8"))
    message.attach(alternative)

    export_path = Path(export_file)
    if export_path.exists():
        with export_path.open("rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={export_path.name}")
        message.attach(part)

    smtp_host = email_cfg["smtp_host"]
    smtp_timeout = int(email_cfg.get("timeout_seconds", 30))
    max_retries = int(email_cfg.get("retry_attempts", 3))
    retry_delay = float(email_cfg.get("retry_delay_seconds", 5))
    use_ssl = email_cfg.get("use_ssl", True)       # port 465 direct SSL (preferred)
    use_tls = email_cfg.get("use_tls", True)       # port 587 STARTTLS (fallback)
    ssl_port = int(email_cfg.get("smtp_ssl_port", 465))
    tls_port = int(email_cfg.get("smtp_port", 587))

    def _try_send(phase_label):
        """Try SSL first (port 465), then STARTTLS (port 587) if SSL fails."""
        # --- Attempt 1: SMTP_SSL on port 465 (immune to STARTTLS interception) ---
        if use_ssl:
            phase = "ssl-connect"
            try:
                with smtplib.SMTP_SSL(smtp_host, ssl_port, timeout=smtp_timeout) as server:
                    phase = "ssl-login"
                    server.login(email_cfg["sender_email"], sender_password)
                    phase = "ssl-sendmail"
                    server.sendmail(email_cfg["sender_email"], recipients, message.as_string())
                logger.info(f"Email sent via SSL (port {ssl_port}) to {', '.join(recipients)} [{phase_label}]")
                return True, None
            except smtplib.SMTPAuthenticationError as exc:
                raise  # Auth errors should not be retried
            except Exception as ssl_exc:
                logger.warning(f"SSL (port {ssl_port}) failed at '{phase}': {ssl_exc}. Trying STARTTLS fallback...")

        # --- Attempt 2: SMTP + STARTTLS on port 587 ---
        if use_tls:
            phase = "starttls-connect"
            with smtplib.SMTP(smtp_host, tls_port, timeout=smtp_timeout) as server:
                phase = "starttls-upgrade"
                server.ehlo()
                server.starttls()
                server.ehlo()
                phase = "starttls-login"
                server.login(email_cfg["sender_email"], sender_password)
                phase = "starttls-sendmail"
                server.sendmail(email_cfg["sender_email"], recipients, message.as_string())
            logger.info(f"Email sent via STARTTLS (port {tls_port}) to {', '.join(recipients)} [{phase_label}]")
            return True, None

        raise RuntimeError("No SMTP method enabled: set use_ssl or use_tls in config")

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            ok, _ = _try_send(f"attempt {attempt}/{max_retries}")
            if ok:
                return True, f"Email sent to {', '.join(recipients)}"
        except smtplib.SMTPAuthenticationError as exc:
            # Auth failures are non-retryable; return explicit guidance (common for Gmail app-password setup).
            logger.warning(f"Email authentication failed: {exc}")
            sender = str(email_cfg.get("sender_email", "")).lower()
            hint = ""
            if "gmail.com" in sender:
                hint = (
                    " Gmail fix: use a Google App Password (not your normal account password), "
                    f"then set env var '{password_env_key}' to that 16-character app password."
                )
            return False, f"Email skipped (auth failed): {exc}.{hint}"
        except Exception as exc:
            last_error = exc
            logger.warning(f"Email attempt {attempt}/{max_retries} failed: {exc}")
            if attempt < max_retries:
                time.sleep(retry_delay * attempt)  # backoff: 5s, 10s, 15s

    logger.error(f"Email delivery failed after {max_retries} attempts. Last error: {last_error}")
    return False, f"Email delivery failed after {max_retries} attempts: {last_error}"

# =============================
# RULE-BASED SCORING ENGINE (REFERENCE)
# =============================

def get_mid_term_horizon(t, score, revenue_growth, earnings_growth, debt_ratio):
    """
    RULE-BASED mid-term goal horizon generator.
    Combines score, trend, growth, and leverage to suggest holding window.
    **NOTE: This is for reference comparison only. AI uses independent scoring.**
    
    Args:
        t (dict): Technical indicators
        score (float): Rule-based score
        revenue_growth (float): Revenue growth %
        earnings_growth (float): Earnings growth %
        debt_ratio (float): Debt-to-equity ratio
        
    Returns:
        str: Suggested mid-term horizon with reason
    """
    bullish = "Bullish" in t.get("macd", "")
    above_200 = "Above 200" in t.get("ema", "")
    rsi = parse_float(t.get("rsi")) or 50

    growth_score = 0
    if revenue_growth is not None:
        growth_score += 1 if revenue_growth > 0 else -1
    if earnings_growth is not None:
        growth_score += 1 if earnings_growth > 0 else -1

    months = 4 + int(score / 12)
    if bullish and above_200:
        months += 2
    if 52 <= rsi <= 68:
        months += 1
    if growth_score > 0:
        months += 1
    if debt_ratio is not None and debt_ratio > 1.5:
        months -= 2

    months = max(3, min(15, months))
    lower = max(3, months - 1)
    upper = min(18, months + 2)

    if growth_score > 0 and bullish and above_200:
        reason = "trend + growth"
    elif bullish or above_200:
        reason = "trend developing"
    elif score >= 45:
        reason = "needs confirmation"
    else:
        reason = "high uncertainty"

    return f"{lower}-{upper} months ({reason})"

def score_logic(t, f, inv):
    """
    RULE-BASED SCORING (Reference only - for comparison with AI).
    Calculates traditional technical + fundamental + investor score (0-60).
    Uses Screener-sourced data for fundamentals (same as new data pipeline).
    **This is kept for comparison purposes. AI uses independent evaluation.**

    Scoring:
      Technical   (max 20): MACD + RSI
      Fundamental (max 20): ROE + ROCE thresholds from Screener
      Investor    (max 20): MF holding + promoter stake quality

    Args:
        t (dict): Technical data
        f (dict): Fundamental data (from get_fundamentals â€” Screener sourced)
        inv (dict): Investor data

    Returns:
        int: Rule-based composite score (0-60)
    """
    score = 0

    # â”€â”€ Technical signals (max 20) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Bullish" in t.get("macd", ""):
        score += 10
    rsi = parse_float(t.get("rsi")) or 50
    if 50 < rsi <= 70:
        score += 10
    elif 40 <= rsi <= 50:
        score += 5

    # â”€â”€ Fundamental signals (max 20) â€” use actual numeric thresholds â”€â”€â”€â”€â”€â”€
    # ROE: Screener value is already in % (e.g. 14.6 = 14.6%)
    roe = parse_float(f.get("roe"))
    if roe is not None:
        if roe > 20:   score += 10
        elif roe > 12: score += 7
        elif roe > 0:  score += 3

    # ROCE: Screener value in %
    roce = parse_float(f.get("roce"))
    if roce is not None:
        if roce > 20:   score += 10
        elif roce > 12: score += 7
        elif roce > 0:  score += 3

    # â”€â”€ Investor signals (max 20) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # MF holding from NSE investor data
    mf = parse_float(inv.get("MF"))
    if mf is not None:
        if mf > 5:   score += 10
        elif mf > 1: score += 5

    # Promoter holding from Screener shareholding table
    promoter = parse_float(f.get("promoter_holding"))
    if promoter is not None:
        if promoter > 50:   score += 10
        elif promoter > 30: score += 5

    return min(60, score)

def decision(score):
    """
    RULE-BASED ACTION (Reference only).
    Maps rule-based score to action recommendation.
    
    Args:
        score (int): Rule-based score
        
    Returns:
        str: Action recommendation
    """
    if score >= 60:
        return "HOLD / ACCUMULATE"
    elif score >= 50:
        return "HOLD"
    else:
        return "WATCHLIST"

# =============================
# AI-BASED RANKING ENGINE (INDEPENDENT)
# =============================



def calculate_ai_score(quote, technical, fundamentals, investor_data, config):
    """
    Sector-aware AI scoring with hard guardrails.
    Keeps the dual-engine design intact, but avoids contradictory BUY calls on weak fundamentals.
    """
    meta = _build_ai_meta(quote, technical, fundamentals, investor_data)
    profile = meta["profile"]
    score = 0.0
    breakdown = {}

    # 1) Valuation (0-15) — sector-aware, PEG de-emphasized because it is unstable when growth is weak/negative.
    valuation_score = 0
    pe_ratio = meta.get("pe_ratio_to_sector")
    pb = meta.get("pb")
    ps = meta.get("ps")
    growth = meta.get("profit_growth_pct")
    if profile["is_financial"]:
        if pb is not None and 0 < pb <= 1.5:
            valuation_score += 10
        elif pb is not None and pb <= 2.5:
            valuation_score += 7
        elif pb is not None and pb <= 4:
            valuation_score += 4
    else:
        if pe_ratio is not None:
            if pe_ratio <= 0.8:
                valuation_score += 8
            elif pe_ratio <= 1.1:
                valuation_score += 6
            elif pe_ratio <= 1.5:
                valuation_score += 3
        if pb is not None and 0 < pb <= 2:
            valuation_score += 4
        elif pb is not None and pb <= 4:
            valuation_score += 2
        if ps is not None and 0 < ps <= 2:
            valuation_score += 3
        elif ps is not None and ps <= 4:
            valuation_score += 1
    if growth is not None and growth < 0:
        valuation_score = max(0, valuation_score - 2)
    valuation_score = min(15, valuation_score)
    breakdown["Valuation (sector-aware)"] = valuation_score
    score += valuation_score

    # 2) Technical setup (0-15) — useful for timing, not enough on its own.
    tech_score = 0
    rsi = parse_float(technical.get("rsi"))
    macd = str(technical.get("macd", ""))
    ema_status = str(technical.get("ema", ""))
    if rsi is not None:
        if 45 <= rsi <= 65:
            tech_score += 7
        elif 40 <= rsi <= 70:
            tech_score += 5
        elif 35 <= rsi < 40 or 70 < rsi <= 75:
            tech_score += 2
    if "Bullish" in macd:
        tech_score += 4
    if "Above 200" in ema_status:
        tech_score += 4
    elif "Above 50" in ema_status:
        tech_score += 2
    breakdown["Technical setup"] = min(15, tech_score)
    score += min(15, tech_score)

    # 3) Quality (0-22)
    quality_score = 0
    roe = meta.get("roe")
    roce = meta.get("roce")
    npm = meta.get("npm_pct")
    opm = meta.get("opm_pct")
    if profile["is_financial"]:
        if roe is not None:
            if roe >= 18:
                quality_score += 12
            elif roe >= 14:
                quality_score += 9
            elif roe >= 10:
                quality_score += 6
            elif roe >= 8:
                quality_score += 3
        if npm is not None:
            if npm >= 18:
                quality_score += 6
            elif npm >= 10:
                quality_score += 4
            elif npm >= 5:
                quality_score += 2
    else:
        if roe is not None:
            if roe >= 20:
                quality_score += 8
            elif roe >= 15:
                quality_score += 6
            elif roe >= 10:
                quality_score += 3
        if roce is not None:
            if roce >= 22:
                quality_score += 10
            elif roce >= 16:
                quality_score += 7
            elif roce >= 12:
                quality_score += 4
        if opm is not None:
            if opm >= 18:
                quality_score += 4
            elif opm >= 12:
                quality_score += 2
    quality_score = min(22, quality_score)
    breakdown["Quality (ROE/ROCE/margins)"] = quality_score
    score += quality_score

    # 4) Growth quality (0-18) — prefer 3Y CAGR, and avoid rewarding unstable/negative growth.
    growth_score = 0
    rev_growth = meta.get("revenue_growth_pct")
    profit_growth = meta.get("profit_growth_pct")
    if rev_growth is not None:
        if rev_growth >= 20:
            growth_score += 8
        elif rev_growth >= 12:
            growth_score += 5
        elif rev_growth >= 5:
            growth_score += 2
    if profit_growth is not None:
        if profit_growth >= 20:
            growth_score += 8
        elif profit_growth >= 12:
            growth_score += 5
        elif profit_growth >= 5:
            growth_score += 2
        elif profit_growth < 0:
            growth_score -= 2
    if rev_growth is not None and profit_growth is not None and abs(rev_growth - profit_growth) <= 12 and rev_growth > 0 and profit_growth > 0:
        growth_score += 2
    growth_score = max(0, min(18, growth_score))
    breakdown["Growth quality"] = growth_score
    score += growth_score

    # 5) Safety / balance sheet (0-18) — sector-aware debt logic.
    safety_score = 0
    debt = meta.get("debt")
    int_cov = meta.get("interest_coverage")
    cash_from_ops = meta.get("cash_from_ops")
    current_ratio = meta.get("current_ratio")
    if profile["debt_metric_applicable"]:
        if debt is not None:
            if debt <= 0.3:
                safety_score += 8
            elif debt <= (1.0 if not profile["debt_soft"] else 2.0):
                safety_score += 5
            elif debt <= (1.8 if not profile["debt_soft"] else 3.0):
                safety_score += 2
        if int_cov is not None:
            if int_cov >= 5:
                safety_score += 5
            elif int_cov >= 3:
                safety_score += 3
            elif int_cov >= 2:
                safety_score += 1
        if current_ratio is not None and current_ratio >= 1.25:
            safety_score += 2
    else:
        # Finance names: debt/equity not economically comparable, so do not punish leverage mechanically.
        safety_score += 4
    if cash_from_ops is not None:
        if cash_from_ops > 0:
            safety_score += 5
        else:
            safety_score -= 4
    safety_score = max(0, min(18, safety_score))
    breakdown["Safety / balance sheet"] = safety_score
    score += safety_score

    # 6) Governance / investability (0-12)
    gov_score = 0
    promoter = meta.get("promoter_holding")
    pledge = meta.get("pledged_pct")
    completeness = meta.get("completeness_ratio")
    if promoter is not None:
        if promoter >= 50:
            gov_score += 4
        elif promoter >= 35:
            gov_score += 2
        elif promoter < 20:
            gov_score -= 1
    if pledge is None or pledge == 0:
        gov_score += 4
    elif pledge <= 5:
        gov_score += 2
    elif pledge > 10:
        gov_score -= 4
    if completeness >= 0.80:
        gov_score += 2
    elif completeness >= 0.60:
        gov_score += 1
    if not meta.get("soft_flags"):
        gov_score += 2
    gov_score = max(0, min(12, gov_score))
    breakdown["Governance / investability"] = gov_score
    score += gov_score

    # Hard guardrails / penalties
    blockers = []
    if meta["risk_flags"]:
        blockers.extend(meta["risk_flags"])
    penalty = 0
    if any("Promoter pledge" in x for x in blockers):
        penalty += 10
    if any("Negative operating cash flow" in x for x in blockers):
        penalty += 10
    if any("Debt stress" in x for x in blockers):
        penalty += 12
    if any("Sub-scale profitability" in x for x in blockers):
        penalty += 10
    if any("Extreme valuation" in x for x in blockers):
        penalty += 8
    if any("Low data completeness" in x for x in blockers):
        penalty += 8
    score = max(0, min(100, round(score - penalty, 1)))

    valuation_view = "Attractive"
    if valuation_score <= 4:
        valuation_view = "Unattractive"
    elif valuation_score <= 7:
        valuation_view = "Stretched"
    elif valuation_score <= 10:
        valuation_view = "Fair"

    meta.update({
        "hard_blockers": blockers,
        "penalty_points": penalty,
        "valuation_view": valuation_view,
        "technical_view": "Bullish" if breakdown["Technical setup"] >= 11 else "Improving" if breakdown["Technical setup"] >= 7 else "Neutral" if breakdown["Technical setup"] >= 4 else "Weak",
        "business_view": "Strong" if quality_score + growth_score >= 28 else "Positive" if quality_score + growth_score >= 20 else "Neutral" if quality_score + growth_score >= 12 else "Weak",
    })
    return {"ai_score": score, "breakdown": breakdown, "meta": meta}



def get_ai_recommendation(ai_score, config, meta=None):
    """Return recommendation after applying sector-aware hard filters."""
    meta = meta or {}
    thresholds = config.get("ai_thresholds", {
        "strong_buy_min_score": 75,
        "buy_min_score": 60,
        "hold_min_score": 45
    })
    blockers = meta.get("hard_blockers", [])
    severe_blockers = [b for b in blockers if any(x in b for x in ["Promoter pledge", "Negative operating cash flow", "Debt stress", "Sub-scale profitability"]) ]
    valuation_view = meta.get("valuation_view", "Fair")
    confidence = "HIGH" if meta.get("completeness_ratio", 0) >= 0.8 and len(blockers) == 0 else "MEDIUM" if meta.get("completeness_ratio", 0) >= 0.6 else "LOW"

    if severe_blockers:
        if ai_score >= thresholds.get("hold_min_score", 45):
            return {
                "recommendation": "HOLD",
                "confidence": confidence,
                "reasoning": "Business quality has offsets, but hard risk blockers mean no fresh buying until they improve.",
                "guardrail": "; ".join(severe_blockers[:3])
            }
        return {
            "recommendation": "REDUCE / AVOID",
            "confidence": confidence,
            "reasoning": "Hard risk blockers dominate the setup; avoid fresh entry until balance-sheet/governance/cash-flow issues improve.",
            "guardrail": "; ".join(severe_blockers[:3])
        }

    if ai_score >= thresholds.get("strong_buy_min_score", 75) and valuation_view in {"Attractive", "Fair"}:
        return {
            "recommendation": "STRONG BUY",
            "confidence": confidence,
            "reasoning": "Balanced quality, growth, safety, and technicals with no hard risk blockers.",
            "guardrail": "Breaks if profitability, cash flow, or governance weakens materially."
        }
    if ai_score >= thresholds.get("buy_min_score", 60):
        if valuation_view == "Unattractive":
            return {
                "recommendation": "HOLD",
                "confidence": confidence,
                "reasoning": "Business/technical setup is constructive, but valuation leaves little room for error right now.",
                "guardrail": "Upgrade only after valuation cools or earnings catch up."
            }
        return {
            "recommendation": "BUY",
            "confidence": confidence,
            "reasoning": "Selective opportunity: fundamentals are acceptable and timing signals are supportive without major red flags.",
            "guardrail": "Downgrade if cash flow turns negative, leverage rises, or the trend breaks."
        }
    if ai_score >= thresholds.get("hold_min_score", 45):
        return {
            "recommendation": "HOLD",
            "confidence": confidence,
            "reasoning": "Mixed setup: enough quality to track, but not enough edge for aggressive fresh buying.",
            "guardrail": "Needs either better valuation or stronger quality/growth confirmation."
        }
    return {
        "recommendation": "REDUCE / AVOID",
        "confidence": confidence,
        "reasoning": "Weak conviction after adjusting for sector context and risk controls.",
        "guardrail": "Revisit only if fundamentals and risk flags improve."
    }



def get_ai_justification(symbol, breakdown, ai_recommendation, quote, technical, config, fundamentals=None, investor_data=None, meta=None):
    """Generate contradiction-free justification text for the AI view."""
    fundamentals = fundamentals or {}
    investor_data = investor_data or {}
    meta = meta or _build_ai_meta(quote, technical, fundamentals, investor_data)
    top_scores = sorted(breakdown.items(), key=lambda x: x[1], reverse=True)[:2]
    top_reasons = [f"{name} ({score} pts)" for name, score in top_scores if score > 0]
    if not top_reasons:
        top_reasons = ["limited scoring evidence"]

    valuation_bits = []
    if meta.get("pe") is not None and meta.get("sector_median_pe"):
        valuation_bits.append(f"PE {round(meta['pe'], 2)}x vs sector median {round(meta['sector_median_pe'], 2)}x")
    if meta.get("pb") is not None:
        valuation_bits.append(f"PB {round(meta['pb'], 2)}x")
    valuation_txt = ", ".join(valuation_bits) if valuation_bits else "valuation data limited"

    risk_flags = meta.get("hard_blockers") or meta.get("soft_flags") or []
    caution = risk_flags[0] if risk_flags else "No major hard blockers detected"
    sector_model = meta.get("profile", {}).get("label", "General")
    default_justification = (
        f"AI Recommendation: {ai_recommendation['recommendation']} ({ai_recommendation['confidence']} confidence). "
        f"Sector model used: {sector_model}. Strongest pillars: {', '.join(top_reasons)}. "
        f"Valuation view is {meta.get('valuation_view', 'Fair')} based on {valuation_txt}. "
        f"Technical view: {meta.get('technical_view', 'Neutral')} with RSI={technical.get('rsi', 'N/A')} and MACD={technical.get('macd', 'N/A')}. "
        f"Main caution: {caution}. {ai_recommendation.get('guardrail', ai_recommendation.get('reasoning', ''))}"
    )

    genai_cfg = config.get("genai", {})
    if not genai_cfg.get("enabled", False):
        logger.info(f"AI justification fallback path used for {symbol}: GenAI disabled in config")
        return default_justification

    system_prompt = (
        "Write a concise investor-friendly stock justification in 3 sentences. "
        "Use only provided data. Mention strengths, one caution, and conclude with the recommendation. "
        "Do not contradict the valuation or risk view."
    )
    user_prompt = json.dumps({
        "recommendation": ai_recommendation.get("recommendation"),
        "confidence": ai_recommendation.get("confidence"),
        "reasoning": ai_recommendation.get("reasoning"),
        "guardrail": ai_recommendation.get("guardrail"),
        "top_scores": top_scores,
        "valuation_view": meta.get("valuation_view"),
        "valuation_text": valuation_txt,
        "technical_view": meta.get("technical_view"),
        "rsi": technical.get("rsi"),
        "macd": technical.get("macd"),
        "risk_flags": risk_flags,
        "sector_model": sector_model,
    }, ensure_ascii=False)
    text = call_genai_ghcp(config, system_prompt, user_prompt)
    return text.strip() if text else default_justification

def analyze(symbol, config):
    """
    MAIN ANALYZER: Dual-layer analysis (Rule-based + AI-based).
    Keeps both systems completely independent for fair comparison.
    
    **Key design: AI analysis is NOT influenced by rule-based results.**
    
    Args:
        symbol (str): Stock symbol
        config (dict): Configuration
        
    Returns:
        dict: Complete analysis row with all metrics, or None if insufficient data
    """
    t = get_technical(symbol)
    if t is None:
        logger.warning(f"Skipping {symbol} - insufficient technical data")
        return None

    f = get_fundamentals(symbol)
    inv = get_investor_data(symbol)
    holder_details = get_holder_details(symbol)

    name, sector = get_company_info(symbol)
    
    # RULE-BASED SCORING (for reference comparison)
    rule_score = score_logic(t, f, inv)

    quote = get_quote_data(symbol)

    # â”€â”€ Override yfinance with Screener Indian-native values where available â”€â”€
    # Screener computes all ratios from INR-denominated financials, eliminating
    # the USD/INR currency-mismatch that inflates yfinance P/S for dual-listed stocks.

    # Profit Margin: Screener NPM% = Net Profit / Revenue * 100 (all INR)
    _npm = parse_float(f.get("npm"))
    if _npm is not None:
        quote["profit_margin"] = _npm / 100           # store as decimal, same as yfinance convention

    # Operating Margin: Screener OPM% string e.g. "13%" â†’ decimal 0.13
    _opm_str = f.get("opm")
    _opm_val = parse_float(str(_opm_str).replace("%", "").strip()) if _opm_str else None
    if _opm_val is not None:
        quote["operating_margin"] = _opm_val / 100

    # P/B: Screener Current Price / Book Value (both INR per share) â†’ no FX risk
    _pb_screener = parse_float(f.get("price_to_book_screener"))
    if _pb_screener is not None:
        quote["price_to_book"] = _pb_screener

    # P/S: Screener MCap (Cr) / Revenue (Cr) â€” pure INR, no currency mismatch at all
    _ps_screener = parse_float(f.get("price_to_sales_screener"))
    if _ps_screener is not None:
        quote["price_to_sales"] = _ps_screener

    # EPS: Screener annual EPS in INR â†’ overrides yfinance (which can reflect USD for ADRs)
    _eps_screener = parse_float(str(f.get("eps", "")).replace(",", ""))
    if _eps_screener is not None:
        quote["eps_trailing_twelve_months"] = _eps_screener

    # Book Value: Screener (INR/share) â†’ overrides yfinance
    _bv_screener = parse_float(str(f.get("book_value", "")).replace(",", ""))
    if _bv_screener is not None:
        quote["book_value"] = _bv_screener

    # Institutional holding: supplement with Screener FII+DII if yfinance missing
    _fii_sc = parse_float(f.get("fii_pct"))
    _dii_sc = parse_float(f.get("dii_pct"))
    if _fii_sc is not None and quote.get("held_percent_institutions") is None:
        quote["held_percent_institutions"] = _fii_sc / 100   # convert % to decimal

    revenue_growth = quote["revenue_growth"]   # yfinance TTM decimal, used as fallback
    earnings_growth = quote["earnings_growth"] # yfinance TTM decimal, used as fallback

    # â”€â”€ Growth: prefer Screener 3Y CAGR over yfinance 1-year TTM â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Screener values come back as plain numbers already in % (e.g. 18.5 = 18.5 %).
    _sg3 = parse_float(f.get("sales_growth_3y"))
    _pg3 = parse_float(f.get("profit_growth_3y"))

    if _sg3 is not None:
        rev_growth_display = f"{round(_sg3, 2)}%"
        rev_growth_source  = "3Y CAGR"
    elif revenue_growth is not None:
        rev_growth_display = f"{round(revenue_growth * 100, 2)}%"
        rev_growth_source  = "1Y TTM"
    else:
        rev_growth_display = "Data unavailable"
        rev_growth_source  = ""

    if _pg3 is not None:
        profit_growth_display = f"{round(_pg3, 2)}%"
        profit_growth_source  = "3Y CAGR"
    elif earnings_growth is not None:
        profit_growth_display = f"{round(earnings_growth * 100, 2)}%"
        profit_growth_source  = "1Y TTM"
    else:
        profit_growth_display = "Data unavailable"
        profit_growth_source  = ""

    # ===== HOLDING INFORMATION =====
    funds_txt = " , ".join([h.replace("(", "= ").replace(")", "") for h in holder_details["fund_houses"]]) if holder_details["fund_houses"] else None
    if holder_details["individuals"]:
        individual_stake = normalize_pct(quote.get("held_percent_insiders"))
        if individual_stake is not None and len(holder_details["individuals"]) > 0:
            per_name = round(individual_stake / len(holder_details["individuals"]), 2)
            individuals_txt = " , ".join([f"{name} = {per_name}%" for name in holder_details["individuals"]])
        else:
            individuals_txt = " , ".join([f"{name} = stake not disclosed" for name in holder_details["individuals"]])
    else:
        individuals_txt = None

    if funds_txt or individuals_txt:
        holding_funds = f"Fund House : {funds_txt or 'Data unavailable from source'} | Individual Investor : {individuals_txt or 'Data unavailable from source'}"
    else:
        holding_funds = "Data unavailable"

    # â”€â”€ MF Holding %: use Screener FII % as a proxy for institutional MF holding â”€â”€
    # Old "Avg. Holding %" was incorrectly averaging top-5 fund stake (~40-50%) which
    # is not the actual MF holding %. Screener's FII% = Foreign Institutional Investors.
    # We show FII% and DII% separately and clearly labelled.
    _fii_pct = parse_float(f.get("fii_pct"))
    _dii_pct = parse_float(f.get("dii_pct"))
    _promo_pct = parse_float(f.get("promoter_holding"))
    # NSE investor data fallback for MF
    _mf_nse = parse_float(inv.get("MF"))

    mf_holding_display = (
    f"{round(_mf_nse, 2)}%" if _mf_nse is not None
    else "Data unavailable"
)
    fii_display = f"{round(_fii_pct, 2)}%" if _fii_pct is not None else "Data unavailable"
    dii_display = f"{round(_dii_pct, 2)}%" if _dii_pct is not None else "Data unavailable"
    promoter_display = f"{round(_promo_pct, 2)}%" if _promo_pct is not None else "Data unavailable"

    # â”€â”€ Debt/Equity: Screener primary, yfinance (already /100 normalised) fallback â”€â”€
    # Sectors where high D/E is structurally normal â€” annotate rather than penalise.
    _high_de_sectors = {
        # Finance / lending
        "banks", "bank", "private banks", "public sector bank", "nbfc",
        "financial services", "insurance", "housing finance",
        "diversified financials", "capital markets", "credit services",
        # Capital-intensive: infrastructure, utilities, real estate
        "utilities", "electric utilities", "gas utilities",
        "infrastructure", "construction", "engineering & construction",
        "real estate", "real estate - development", "reit",
        # Healthcare â€” hospitals carry high capex-funded debt
        "medical care facilities", "hospitals", "healthcare facilities",
        "hospital", "healthcare",
    }
    _industry_lower = str(quote.get("industry", "")).lower()
    _sector_lower   = str(quote.get("sector_yf", "")).lower()
    _is_high_de_sector = (
        any(s in _industry_lower for s in _high_de_sectors)
        or any(s in _sector_lower  for s in _high_de_sectors)
        or "financ" in _industry_lower or "financ" in _sector_lower
        or "bank"   in _industry_lower or "bank"   in _sector_lower
        or "real estate" in _industry_lower
        or "hospital"   in _industry_lower
    )
    _is_banking = _is_high_de_sector   # keep alias for backward compatibility

    debt_value = clean(f.get("debt"))
    if debt_value == "Data unavailable" and quote["debt_to_equity"] is not None:
        debt_value = round(quote["debt_to_equity"], 2)
    if _is_banking and debt_value != "Data unavailable":
        debt_value = f"{debt_value} (high leverage expected for banks)"

    # â”€â”€ Dividend Yield: Screener primary, yfinance fallback â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Screener returns dividend yield as a number already in % (e.g. 2.5 = 2.5%).
    # yfinance dividendYield is a decimal (0.025 = 2.5%); guard against rare cases
    # where Yahoo returns it already as a percentage (> 0.5 threshold).
    _div_screener = parse_float(f.get("dividend_yield"))
    _div_yf_raw   = quote.get("dividend_yield")  # normalised decimal from get_quote_data
    if _div_screener is not None:
        dividend_display = f"{round(_div_screener, 2)}%"
    elif _div_yf_raw is not None:
        # Guard: if yfinance somehow returned it already as % (> 0.5), don't double-multiply
        _div_pct = _div_yf_raw if _div_yf_raw > 0.5 else _div_yf_raw * 100
        dividend_display = f"{round(_div_pct, 2)}%"
    else:
        dividend_display = "Data unavailable"

    mid_term_horizon = get_mid_term_horizon(t, rule_score, revenue_growth, earnings_growth, quote["debt_to_equity"])

    # ===== RULE-BASED RECOMMENDATION (for reference) =====
    rule_recommendation = decision(rule_score)

    # ===== AI-BASED ANALYSIS (COMPLETELY INDEPENDENT) =====
    if config.get("enable_ai_ranking", True):
        ai_data = calculate_ai_score(quote, t, f, inv, config)
        ai_score = ai_data["ai_score"]
        ai_breakdown = ai_data["breakdown"]
        ai_meta = ai_data.get("meta", {})
        ai_rec = get_ai_recommendation(ai_score, config, ai_meta)
        ai_justification = get_ai_justification(symbol, ai_breakdown, ai_rec, quote, t, config, f, inv, ai_meta)
    else:
        ai_score = "N/A"
        ai_breakdown = {}
        ai_meta = {"profile": {"label": "General"}, "data_quality_label": "LOW", "hard_blockers": []}
        ai_rec = {"recommendation": "N/A", "confidence": "N/A", "guardrail": "AI ranking disabled in config"}
        ai_justification = "AI ranking disabled in config"

    # â”€â”€ PEG Ratio (Screener-computed: PE / 3Y profit CAGR) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# PEG Ratio display
    _peg_computed = parse_float(f.get("peg_computed"))
    _peg_yf = quote.get("peg_ratio")

    if _peg_computed is not None and _pg3 is not None and _pg3 > 0 and _pg3 <= 60:
        peg_display = f"{round(_peg_computed, 2)} (PE/{_pg3}% profit CAGR)"
    elif _peg_yf is not None and _pg3 is not None and _pg3 > 0 and _pg3 <= 60 and _peg_yf < 10:
        peg_display = f"{round(_peg_yf, 2)} (yfinance, low confidence)"
    elif _pg3 is not None and _pg3 <= 0:
        peg_display = "Not meaningful (profit growth <= 0)"
    elif _pg3 is not None and _pg3 > 60:
        peg_display = "Not meaningful (growth too volatile/high)"
    else:
        peg_display = "Data unavailable"

    # Sector PE Comparison
    _stock_pe = parse_float(f.get("pe"))
    _scr_sector = f.get("screener_sector") or sector or quote.get("sector_yf") or ""
    _scr_indust = f.get("screener_industry") or quote.get("industry") or ""
    _median_pe, _pe_label = get_sector_pe(_scr_sector, _scr_indust)

    if _stock_pe is not None and _stock_pe > 0 and _median_pe:
        _pe_ratio = _stock_pe / _median_pe

        if _pe_ratio < 0.8:
            _pe_verdict = f"CHEAP ({round(_stock_pe, 2)}x vs sector median {round(_median_pe, 2)}x)"
        elif _pe_ratio < 1.2:
            _pe_verdict = f"FAIR ({round(_stock_pe, 2)}x vs sector median {round(_median_pe, 2)}x)"
        else:
            _pe_verdict = f"PRICEY ({round(_stock_pe, 2)}x vs sector median {round(_median_pe, 2)}x)"

        sector_pe_display = sanitize_text(f"{_pe_verdict} [{_pe_label}]")
    else:
        sector_pe_display = "Data unavailable"

    # last derived display variables above ...
    # sector_pe_display = ...

    # FAIL-SAFE: ensure inst_trend_display always exists
    if "inst_trend_display" not in locals():
        inst_trend_display = "Data unavailable"
# Final Rating
    if "final_rating" not in locals():
        final_rating = get_final_rating(
            ai_score=ai_score if isinstance(ai_score, (int, float)) else 0,
            rule_score=rule_score,
            roe=f.get("roe"),
            roce=f.get("roce"),
            debt_ratio=quote.get("debt_to_equity"),
            profit_growth_3y=f.get("profit_growth_3y"),
            revenue_growth_3y=f.get("sales_growth_3y"),
            promoter_holding=f.get("promoter_holding"),
            is_high_de_sector=get_sector_profile(
                f.get("screener_sector") or sector or quote.get("sector_yf") or "",
                f.get("screener_industry") or quote.get("industry") or ""
            ).get("is_bank_like", False),
        )
    final_rating = align_composite_with_ai(final_rating, ai_rec.get("recommendation"))
    
    # Debt / Equity display (sector-aware and Excel-safe)    # Debt / Equity display (sector-aware and Excelustry") or quote.get("industry") or ""
    _profile = get_sector_profile(_scr_sector, _scr_indust)
    _is_bank_like = _profile.get("is_bank_like", False)

    combined_sector_text = f"{_scr_sector} {_scr_indust}".lower()

    _debt_val = parse_float(f.get("debt"))
    if _debt_val is None:
        _debt_val = quote.get("debt_to_equity")

    if _debt_val is not None:
        if "insurance" in combined_sector_text:
            debt_value = f"{round(_debt_val, 2)} (insurance / financial balance-sheet model)"
        elif any(tok in combined_sector_text for tok in ["exchange", "capital markets", "depository", "asset management", "market infrastructure"]):
            debt_value = f"{round(_debt_val, 2)} (financial infrastructure balance-sheet model)"
        elif _is_bank_like:
            debt_value = f"{round(_debt_val, 2)} (banking / lending leverage model)"
        else:
            debt_value = f"{round(_debt_val, 2)}"
    else:
        debt_value = "Data unavailable"
    _scr_sector = f.get("screener_sector") or sector or ""


    print("[DEBUG] final_rating exists:", "final_rating" in locals(), locals().get("final_rating"))

# ===== OWNERSHIP FALLBACK FIX =====
    # Rebuild ownership displays safely just before return, so final row always has values.

    _fii_pct = parse_float(inv.get("FII"))
    if _fii_pct is None:
        _fii_pct = parse_float(f.get("fii_pct"))

    _dii_pct = parse_float(inv.get("DII"))
    if _dii_pct is None:
        _dii_pct = parse_float(f.get("dii_pct"))

    _promo_pct = parse_float(f.get("promoter_holding"))
    if _promo_pct is None:
        _promo_pct = parse_float(inv.get("PromoterHolding"))

    _mf_nse = parse_float(inv.get("MF"))
    if _mf_nse is None:
        _mf_nse = parse_float(f.get("mf_pct"))
    if _mf_nse is None:
        _mf_nse = parse_float(f.get("mutual_funds"))
    if _mf_nse is None:
        _mf_nse = parse_float(f.get("mutualFund"))
    if _mf_nse is None:
        _mf_nse = parse_float(f.get("mutualFunds"))

    fii_display = f"{round(_fii_pct, 2)}%" if _fii_pct is not None else "Data unavailable"
    dii_display = f"{round(_dii_pct, 2)}%" if _dii_pct is not None else "Data unavailable"
    promoter_display = f"{round(_promo_pct, 2)}%" if _promo_pct is not None else "Data unavailable"

    # MF must ONLY use true MF data. Do not proxy FII into MF.
    mf_holding_display = f"{round(_mf_nse, 2)}%" if _mf_nse is not None else "Data unavailable"

    # FII / DII trend: if 3Q trend text is unavailable, show latest current values.
    inst_trend_display = "Data unavailable"

    _fii_trend_raw = f.get("fii_trend")
    _dii_trend_raw = f.get("dii_trend")

    if _fii_trend_raw and _dii_trend_raw:
        fii_trend_text = re.sub(r"\s*>\s*", " -> ", str(_fii_trend_raw))
        dii_trend_text = re.sub(r"\s*>\s*", " -> ", str(_dii_trend_raw))
        inst_trend_display = sanitize_text(
            f"FII: {fii_trend_text} | DII: {dii_trend_text}"
        )
    elif (_fii_pct is not None) or (_dii_pct is not None):
        inst_trend_display = sanitize_text(
            f"Latest FII: {fii_display} | DII: {dii_display}"
        )


    return {
        # ===== IDENTIFICATION =====
        "Company Name": symbol.replace(".NS", ""),
        "Company Full Name": sanitize_text(name),
        "Sector": sanitize_text(sector),
        "Market Cap Category": classify_mcap(f.get("mcap")),

        # ===== HOLDINGS & OWNERSHIP =====
        "Holding Funds (Top 5)": sanitize_text(holding_funds),
        "Promoter Holding (%)": promoter_display,
        "FII Holding (%)": fii_display,
        "DII Holding (%)": dii_display,
        "MF Holding (%)": mf_holding_display,
        "FII / DII Trend (3Q)": sanitize_text(inst_trend_display),

        # ===== FUNDAMENTAL METRICS =====
        "PE Ratio": clean(f.get("pe")),
        "ROE (%)": clean(f.get("roe")),
        "ROCE (%)": clean(f.get("roce")),
        "Debt/ Equity": sanitize_text(debt_value),
        "Rev. Growth 3Y (%)": rev_growth_display,
        "Profit Growth 3Y (%)": profit_growth_display,

        # ===== VALUATION METRICS =====
        "Price-to-Book": format_ratio(quote.get("price_to_book")),
        "Price-to-Sales": format_ratio(quote.get("price_to_sales")),
        "PEG Ratio (computed)": sanitize_text(peg_display),
        "Sector PE Comparison": sanitize_text(sector_pe_display),
        "Beta": round(quote["beta"], 2) if quote.get("beta") is not None else "Data unavailable",

        # ===== QUALITY METRICS =====
        "Dividend Yield (%)": dividend_display,
        "Profit Margin (%)": f"{round(quote['profit_margin'] * 100, 2)}%" if quote.get("profit_margin") is not None else "Data unavailable",
        "Operating Margin (%)": f"{round(quote['operating_margin'] * 100, 2)}%" if quote.get("operating_margin") is not None else "Data unavailable",

        # ===== TECHNICAL METRICS =====
        "Current Price (INR)": t["price"],
        "52W High (INR)": t["high"],
        "52W Low (INR)": t["low"],
        "Price vs EMA (20/50/200)": sanitize_text(t["ema"]),
        "RSI (14)": t["rsi"],
        "MACD Signal": sanitize_text(t["macd"]),

        # ===== RULE-BASED ANALYSIS =====
        "Rule-Based Score": rule_score,
        "Short-Term Goal Fit": "Momentum trade" if rule_score > 55 else "Pullback entry",
        "Mid-Term Goal Horizon": sanitize_text(mid_term_horizon),
        "Rule-Based Rating": sanitize_text(rule_recommendation),

        # ===== AI-BASED ANALYSIS =====
        "AI Score": ai_score,
        "AI Ranking": sanitize_text(f"Rank {ai_rec['recommendation']} ({ai_rec['confidence']})"),
        "AI Recommendation": sanitize_text(ai_rec["recommendation"]),
        "AI Confidence": sanitize_text(ai_rec["confidence"]),
        "AI Justification": sanitize_text(ai_justification),
        "Sector Model": sanitize_text(ai_meta.get("profile", {}).get("label", "General")),
        "Data Quality": sanitize_text(ai_meta.get("data_quality_label", "LOW")),
        "Risk Flags": non_empty_flags(
            (ai_meta.get("hard_blockers", []) or ai_meta.get("soft_flags", []))
            if isinstance(ai_meta, dict) else []
        ),
        "Why AI View Can Be Wrong": sanitize_text(
            ai_rec.get("guardrail", "Monitor balance-sheet, valuation, and trend changes")
        ),

        # ===== FINAL UNIFIED RATING =====
        "Composite Rating": sanitize_text(final_rating),
        "Final Rating": sanitize_text(final_rating),  # backward compatibility
        # ===== METADATA =====
        "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
# =============================
# MAIN EXECUTION
# =============================

def main():
    """
    MAIN ENTRY POINT: Load config, process all stocks, generate output.

        Supports two stock selection modes:
            1. Predefined Symbols  â€” symbols listed in config.json
            2. AI Live Selection   â€” discover top symbols using rule keys in skill.md

    Workflow:
    1. Parse CLI arguments and load config
    2. Resolve stock symbols (predefined or live AI selection)
    3. Process each symbol: technical, fundamental, investor data
    4. Calculate BOTH rule-based (reference) and AI-based (primary) recommendations
    5. Export to Excel, write run report, archive snapshot, send email
    """
    parser = argparse.ArgumentParser(
        description="AI-Powered Stock Analysis Agent",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Modes:\n"
            "  predefined   Use stock symbols defined in config.json\n"
            "  skill        Use live AI selection with rule keys from skill.md\n"
            "\nExamples:\n"
            "  python Stock_Agent.py                          # use mode from config.json\n"
            "  python Stock_Agent.py --mode predefined\n"
            "  python Stock_Agent.py --mode skill\n"
            "  python Stock_Agent.py --mode skill --skill-file MY_RULES.md\n"
        ),
    )
    parser.add_argument(
        "--mode",
        choices=["predefined", "skill"],
        default=None,
        help="'predefined' = symbols from config.json; 'skill' = live AI selection using skill.md rules",
    )
    parser.add_argument(
        "--skill-file",
        default=None,
        metavar="FILE",
        help="Path to rules markdown file for live selection (overrides symbol_selection.live_rules_file in config.json)",
    )
    parser.add_argument(
        "--config",
        default="config.json",
        metavar="FILE",
        help="Path to config JSON file (default: config.json)",
    )
    parser.add_argument(
        "--ask-stock",
        default=None,
        metavar="STOCK",
        help="Answer a natural-language question for one stock using the generated Excel report.",
    )
    parser.add_argument(
        "--question",
        default=None,
        metavar="TEXT",
        help="Natural-language question used with --ask-stock.",
    )
    parser.add_argument(
        "--excel-file",
        default=None,
        metavar="FILE",
        help="Excel report file to query (defaults to config export_file).",
    )
    args = parser.parse_args()

    logger.info("=" * 80)
    logger.info("Starting AI Stock Analysis Agent")
    logger.info("=" * 80)

    run_started_at = datetime.now()
    config = load_config(args.config)

    # Optional CLI override: keep backward compatibility for --mode skill/predefined
    if args.mode:
        config.setdefault("symbol_selection", {})
        config["symbol_selection"]["mode"] = "live_refresh" if args.mode == "skill" else "predefined"
    if args.skill_file:
        config.setdefault("symbol_selection", {})
        config["symbol_selection"]["live_rules_file"] = args.skill_file
    if args.ask_stock:
        if not args.question:
            parser.error("--question is required when using --ask-stock")
        response = answer_stock_question_from_report(args.ask_stock, args.question, config, args.excel_file)
        print(json.dumps(response, indent=2, ensure_ascii=False, default=str))
        return

    cache_status = get_ai_cache_status_line(config)
    logger.info(cache_status)
    print(f"[INFO] {cache_status}")

    stocks, symbol_source = select_symbols_for_run(config)
    mode_label = {
        "PREDEFINED": "PREDEFINED SYMBOLS",
        "LIVE_CACHED": "AI SUGGESTED SYMBOLS (CACHED)",
        "LIVE_REFRESH": "AI LIVE REFRESHED SYMBOLS",
    }.get(symbol_source, "PREDEFINED SYMBOLS")
    logger.info(f"Analysis mode: {mode_label}")

    # Populate query scores for all symbols regardless of source mode
    logger.info("Populating query scores for Excel transparency...")
    populate_query_scores_for_symbols(stocks, config)

    logger.info(f"Processing {len(stocks)} stocks [{mode_label}]")
    print(f"\n[INFO] Mode: {mode_label}")
    print(f"[INFO] Stocks to analyze: {len(stocks)}")

    results = []
    for i, symbol in enumerate(stocks, 1):
        print(f"[{i}/{len(stocks)}] Processing {symbol}...")
        logger.info(f"Processing {symbol}")
        r = analyze(symbol, config)
        if r:
            results.append(r)
        time.sleep(0.5)  # Rate limiting to avoid API throttling

    if not results:
        logger.error("No valid results generated")
        print("âŒ No valid results generated")
        return

    df = pd.DataFrame(results)

    # Sort by AI Score (descending) if available
    try:
        if "AI Score" in df.columns:
            df = add_score_helpers(df)
            df = df.sort_values(by="AI Score Numeric", ascending=False)
            df = df.drop(["AI Score Numeric", "Rule Score Numeric"], axis=1)
    except Exception:
        pass

    df["Rank"] = range(1, len(df) + 1)

    # If helper columns exist, refresh Rank Numeric
    if "Rank Numeric" in df.columns:
        df["Rank Numeric"] = pd.to_numeric(df["Rank"], errors="coerce").fillna(99999)

    # Drop helper columns before export if you don't want them in final sheet
    for helper_col in ["AI Score Numeric", "Rule Score Numeric", "Rank Numeric"]:
        if helper_col in df.columns:
            df.drop(columns=[helper_col], inplace=True)

    # Add query score transparency columns
    df = add_query_score_columns(df, config)

    export_file = config.get("export_file", "AI_STOCK_ANALYSIS.xlsx")
    comparison_df, comparison_label = load_comparison_snapshot(
        config.get("history_dir", "analysis_history"), run_started_at
    )
    print("[DEBUG] DataFrame columns:", df.columns.tolist())
    report_sections = build_run_report(df, comparison_df, comparison_label, config)
    report_sections["mode_label"] = mode_label   # surface mode in all report outputs

    justification_df = build_score_justification_sheet(df)
    export_path = Path(export_file)
    export_path.parent.mkdir(parents=True, exist_ok=True)
    # Final text sanitization before export
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(sanitize_text)

    # Remove duplicate Final Rating once Composite Rating exists
    if "Composite Rating" in df.columns and "Final Rating" in df.columns:
        df = df.drop(columns=["Final Rating"])
    with tempfile.NamedTemporaryFile(delete=False, suffix=export_path.suffix or ".xlsx", dir=str(export_path.parent)) as tmp:
        tmp_export = tmp.name
    with pd.ExcelWriter(tmp_export, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Analysis", index=False)
        justification_df.to_excel(writer, sheet_name="Score_Justification", index=False)
        apply_score_justification_colors(writer, config)
    Path(tmp_export).replace(export_path)
    export_file = str(export_path)
    write_run_report(
        config.get("run_report_file", "latest_run_report.txt"),
        report_sections,
        export_file,
        len(df),
        run_started_at,
    )
    archive_path = archive_analysis_snapshot(
        df, config.get("history_dir", "analysis_history"), run_started_at
    )
    _, email_status = send_email_report(
        config, report_sections, export_file, len(df), run_started_at, mode_label
    )

    logger.info(f"Analysis complete. Results written to {export_file}")
    logger.info(f"Total stocks analyzed: {len(df)}")
    logger.info(f"Archived snapshot: {archive_path}")
    logger.info(email_status)
    print(f"\n[OK] Analysis complete. Output: {export_file}")
    print(f"[INFO] Mode: {mode_label}")
    print(f"[INFO] Total stocks analyzed: {len(df)}")
    print(f"[INFO] Snapshot archived to: {archive_path}")
    print(f"[INFO] Report file: {config.get('run_report_file', 'latest_run_report.txt')}")
    print(f"[INFO] Email status: {email_status}")
    print("\n" + "=" * 120)
    print("TOP 10 STOCKS BY AI SCORE (with comparison to Rule-Based):")
    print("=" * 120)
    print(df[["Company Name", "AI Score", "AI Recommendation", "Rule-Based Rating", "AI Confidence"]].head(10).to_string(index=False))
    print("\nWEEKLY SUMMARY")
    print("=" * 120)
    print(report_sections["weekly_summary"])

if __name__ == "__main__":
    main()
